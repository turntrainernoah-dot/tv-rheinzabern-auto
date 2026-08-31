#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TV Rheinzabern – Automatischer Trainingsplan Generator
=======================================================
Läuft via GitHub Actions alle 30 Minuten.
PC muss NICHT an sein.

Ablauf:
  check_quick.py → needs_update=true → DIESER SCRIPT läuft
  - Plan vorhanden + Abwesenheitsänderung → Plan aktualisieren (Trainer unverändert)
  - Kein Plan + Veröffentlichungsfenster (Mi/Fr 22:00 CEST) → Neuen Plan erstellen
  - Trainer-Wechsel → E-Mail, manuell prüfen
"""

import json, hashlib, os, re, sys, tempfile, shutil, subprocess, itertools, urllib.request, urllib.parse
from datetime import date, timedelta, datetime, timezone

import paramiko
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════
#  STATISCHE KONFIGURATION
# ════════════════════════════════════════════════════════════════

# Phase 2 (24.08.2026): aus dem oeffentlichen Repo ausgelagert. ALLE_TURNER/
# ALLE_TRAINER/WEBSITE_TO_DISPLAY enthielten bisher die echten Klarnamen aller
# Kinder und Trainer hartkodiert im oeffentlichen Repo. Sie sind jetzt LEER
# und werden ausschliesslich zur Laufzeit aus dem geschuetzten Server-Ordner
# geladen (apply_config_roster() unten, unveraendert seit Stage 4 des
# Admin-Bereich-Systems -- neu ist nur, dass es keinen Namens-Fallback mehr
# gibt). Kann config/config.json nicht geladen werden, bricht main() klar ab
# (REQUIRE_SERVER_ROSTER-Check), statt mit leerem Roster einen kaputten oder
# veralteten Plan zu erzeugen.
ALLE_TURNER = {}
ALLE_TRAINER = []

# Mapping: Website-Format (nach normalize) → Anzeigename -- wird aus
# config/config.json befuellt (apply_config_roster).
WEBSITE_TO_DISPLAY = {}

# Reihenfolge der konfigurierten Gruppen (z.B. ["G1","G2","G3"], 3-6 Gruppen
# moeglich) -- wird aus config/config.json befuellt (apply_config_roster).
GRUPPEN_ORDER = []

# Pro Gruppe/Wochentag die drei Phasen-Zeitfenster (Minuten seit 0:00):
# {gruppe: {"mi": {"aufwaermen":{"start":m,"ende":m}, "geraet1":{...}, "geraet2":{...}}, "fr": {...}}}
# wird aus config/config.json befuellt (apply_config_roster).
GRUPPEN_ZEITEN = {}

# Trainer-Namen, die nur eingeteilt werden sollen, wenn es sonst nicht fuer
# alle Gruppen reicht ("immer_springer" in config.json). Wird aus
# config/config.json befuellt (apply_config_roster).
IMMER_SPRINGER = set()

# Gruppennamen, die ihre beiden Geraete-Bloecke in VERTAUSCHTER Reihenfolge
# turnen ("geraet_tausch": true in config.json, admin.php-Checkbox "Geraet-
# Reihenfolge tauschen"): waehrend ihres zeitlich ERSTEN Blocks (Config-
# Schluessel "geraet1") stehen sie tatsaechlich auf dem physischen Geraet 2,
# waehrend des zweiten ("geraet2") auf Geraet 1 -- siehe _effektive_phase().
# Loest dieselbe Kapazitaets-Kollision wie die versetzte Zeit (siehe
# _VERSETZTE_ZEITEN unten), aber ohne die Uhrzeit zu verschieben: 3+ Gruppen
# koennen dadurch zeitgleich trainieren, solange nie mehr als 2 gleichzeitig
# auf demselben physischen Geraet stehen. Wird aus config/config.json
# befuellt (apply_config_roster).
GRUPPEN_TAUSCH = set()

GERAETE_ROTATION = [
    ("Boden", "Barren"),
    ("Sprung", "Reck"),
    ("Seitpferd", "Ringe"),
]

# Migrations-/Fallback-Zeiten (Stand vor dem Gruppenzeiten-Umbau, 25.08.2026):
# G1/G2 Mi=Fr 17:00-19:00 in drei gleich langen Phasen, G3/G4 Mi=Fr eine
# halbe Stunde versetzt (17:00-17:30 kein Training). Greift, wenn eine
# Gruppe in config.json kein "zeiten"-Feld hat (alte Config oder Fehler) --
# siehe apply_config_roster().
#
# G3 bekommt bewusst die VERSETZTE Zeit (wie G4), nicht dieselbe wie G1/G2:
# waeren alle drei (G1/G2/G3) zeitgleich, waeren zu Geraet1-Zeit alle drei
# gleichzeitig auf demselben Geraet -- das verletzt die neue Kapazitaets-
# Grenze "max. 2 Gruppen gleichzeitig pro Geraet" (Noah-Entscheidung, siehe
# admin.php::validateGruppenZeiten()). Die alte Live-Zuteilung kam nie in
# diese Lage, weil G3+G4 vor diesem Umbau ohnehin dauerhaft zusammengelegt
# waren (siehe CLAUDE.md/Vault [[G3+G4 Dauer-Zusammenlegung]]) -- die
# versetzte Default-Zeit fuer G3 setzt dieselbe zeitliche Trennung fort,
# jetzt als eigene Gruppen-Config statt als Code-Sonderfall. G3 und G4 haben
# dadurch identische Default-Zeiten und sind damit sofort wieder
# zeit-kompatibel (zeiten_kompatibel()), falls ein Notfall-Merge noetig wird.
_STANDARD_ZEITEN = {"aufwaermen": {"start": "17:00", "ende": "17:30"},
                     "geraet1":   {"start": "17:30", "ende": "18:15"},
                     "geraet2":   {"start": "18:15", "ende": "19:00"}}
_VERSETZTE_ZEITEN = {"aufwaermen": {"start": "17:30", "ende": "18:15"},
                      "geraet1":   {"start": "18:15", "ende": "19:00"},
                      "geraet2":   {"start": "19:00", "ende": "19:30"}}

def _default_zeiten_for(gruppe_name):
    """Fallback-Zeiten fuer eine Gruppe ohne 'zeiten'-Feld in config.json.
    'G3'/'G4' (Alt-Sonderfall) bekommen die versetzte Zeit, alles andere die
    Standard-Zeit -- siehe _STANDARD_ZEITEN/_VERSETZTE_ZEITEN oben."""
    tag = dict(_VERSETZTE_ZEITEN) if gruppe_name in ("G3", "G4") else dict(_STANDARD_ZEITEN)
    return {"mi": {k: dict(v) for k, v in tag.items()},
            "fr": {k: dict(v) for k, v in tag.items()}}

FARBEN = {
    "g1_blau":        "2471A3",
    "g2_orange":      "CA6F1E",
    "aufwaermen":     "A569BD",
    "aufbauen":       "717D7E",
    "springer":       "A0522D",
    "abwesend":       "D5D8DC",
    "sonder":         "E74C3C",
    "titel":          "2C3E50",
    "header":         "5D6D7E",
    "anwesend":       "1E8449",
    "abwesend_turner":"C0392B",
    "anmerkung":      "D6EAF8",
    "legende_bg":     "2C3E50",
}

def farbe_fuer_phase(phase):
    """Einzige Quelle fuer die Zellfarbe einer Phase (2-Farben-Modell, ersetzt
    das fruehere 4-Farben-Schema): Aufwaermen -> lila, Geraet1 -> blau,
    Geraet2 -> orange, alles andere (kein Training gerade / Luecke) ->
    Aufbauen-Grau. Die max.-2-Gruppen-pro-Geraet-Grenze ist seit dem
    Gruppenzeiten-Umbau eine Config-Invariante (admin.php-Validierung).
    `phase` muss hier bereits die EFFEKTIVE (Tausch-bereinigte) Phase sein --
    siehe _effektive_phase(). Ein globales Alternierungs-Flag wie frueher
    g1_starts_geraet2 ist nicht mehr noetig, weil GRUPPEN_TAUSCH denselben
    Zweck pro Gruppe uebernimmt."""
    return {"aufwaermen": "aufwaermen", "geraet1": "g1_blau", "geraet2": "g2_orange"}.get(phase, "aufbauen")

def _effektive_phase(gruppe, phase):
    """Wendet GRUPPEN_TAUSCH auf eine rohe Config-Phase an: steht `gruppe` auf
    Tausch, wird 'geraet1' zu 'geraet2' und umgekehrt (Aufwaermen bleibt
    unveraendert) -- so turnt die Gruppe zeitlich weiterhin in ihrem
    konfigurierten Block, aber auf dem jeweils ANDEREN physischen Geraet.
    Muss vor jedem farbe_fuer_phase()-Aufruf auf die rohe Phase angewendet
    werden."""
    if phase in ("geraet1", "geraet2") and gruppe in GRUPPEN_TAUSCH:
        return "geraet2" if phase == "geraet1" else "geraet1"
    return phase

# ════════════════════════════════════════════════════════════════
#  UMGEBUNGSVARIABLEN (GitHub Secrets)
# ════════════════════════════════════════════════════════════════

SSH_HOST      = os.environ.get("SSH_HOST",     "access-5017462830.webspace-host.com")
SSH_USER      = os.environ.get("SSH_USER",     "a2358459")
SSH_PASSWORD  = os.environ.get("SSH_PASSWORD", "")
SSH_PORT      = int(os.environ.get("SSH_PORT", "22"))
WHATSAPP_PHONE    = os.environ.get("WHATSAPP_PHONE", "")
CALLMEBOT_APIKEY  = os.environ.get("CALLMEBOT_APIKEY", "")
# Manueller Regenerations-Erzwinger (workflow_dispatch-Input "force" in
# auto_trainingsplan.yml) -- fuer den Fall, dass ein reiner Code-Fix (z.B. an
# der Trainer-Zuteilungslogik) auf einen bereits generierten Plan angewendet
# werden soll, OHNE echte Abwesenheiten/Anmerkungen zu aendern. Bypasst NUR
# die "nichts geaendert"-Kurzschluesse, nie den Schutz fuer protected_plans
# oder ein gesperrtes lock_trainer_plan.
FORCE_REGEN = os.environ.get("FORCE_REGEN", "").lower() == "true"

# ════════════════════════════════════════════════════════════════
#  STATE (gespeichert auf dem Server als state_auto.json)
# ════════════════════════════════════════════════════════════════

DEFAULT_STATE = {
    "last_training_date": "17.06.2026",
    "geraet_combo_index": 1,      # 0=Boden+Barren, 1=Sprung+Reck, 2=Seitpferd+Ringe
    "generated_plans":   ["10.06.26", "17.06.26"],
    "plan_data":         {},
}

def load_state(sftp):
    try:
        f    = sftp.open("state_auto.json", "r")
        data = json.loads(f.read().decode("utf-8"))
        f.close()
        print("State von Server geladen.")
        return data
    except Exception:
        print("Kein State auf Server – verwende Default.")
        return DEFAULT_STATE.copy()

def save_state(sftp, state):
    data = json.dumps(state, indent=2, ensure_ascii=False).encode("utf-8")
    f    = sftp.open("state_auto.json", "wb")
    f.write(data)
    f.close()
    print("State auf Server gespeichert.")

# ════════════════════════════════════════════════════════════════
#  HILFSFUNKTIONEN
# ════════════════════════════════════════════════════════════════

def compute_absences_hash(absences):
    return hashlib.md5(
        json.dumps(absences, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()

def is_publication_window():
    """Veröffentlichungsfenster: Mi oder Fr nach 20:00 UTC (= 22:00 CEST)."""
    now = datetime.now(timezone.utc)
    return now.weekday() in (2, 4) and now.hour >= 20

def get_gear_from_plan_data(state):
    """
    Berechnet den korrekten Geräte-Combo-Index aus plan_data (nicht aus geraet_combo_index).
    Sucht den neuesten Eintrag in plan_data mit Gerätedaten und leitet den nächsten Index ab.
    Gibt zurück: (letzter_combo_idx, g1_starts_geraet2, letztes_datum) oder (None, None, None).
    """
    plan_data = state.get("plan_data", {})
    if not plan_data:
        return None, None, None

    dated_plans = []
    for datum_kurz, pdata in plan_data.items():
        if pdata and "geraet_1" in pdata and "geraet_2" in pdata:
            try:
                d = datetime.strptime(datum_kurz, "%d.%m.%y").date()
                dated_plans.append((d, datum_kurz, pdata))
            except Exception:
                pass

    if not dated_plans:
        return None, None, None

    dated_plans.sort(key=lambda x: x[0])
    latest_date, latest_key, latest = dated_plans[-1]

    g1 = latest.get("geraet_1")
    g2 = latest.get("geraet_2")
    for idx, (r1, r2) in enumerate(GERAETE_ROTATION):
        if r1 == g1 and r2 == g2:
            print(f"[ROTATION] Letzter Plan: {latest_key} = {g1}+{g2} (Combo {idx}) → nächster: {(idx+1)%3}")
            return idx, latest.get("g1_starts_geraet2", False), latest_date

    print(f"[ROTATION-WARN] Kombo {g1}+{g2} nicht in GERAETE_ROTATION gefunden – Fallback auf State.")
    return None, None, None


def get_next_gear(state, exclude_date=None, fixed_entries=None):
    """Bestimmt die Geraete fuers naechste Training aus den letzten (bis zu) 8
    ECHTEN Trainingsplaenen in plan_data. Ausgefallene Trainings
    (entfall_published) und Eintraege ohne gueltige Geraetekombination werden
    ignoriert. Vom juengsten gueltigen Plan wird genau EIN Schritt weitergerueckt.
    Reihenfolge (Zyklus): Sprung+Reck -> Seitpferd+Ringe -> Boden+Barren.
    Gibt combo_idx zurueck. geraet_combo_index wird bewusst NICHT mehr
    verwendet (vermeidet Drift / uebersprungene Geraete). Das fruehere
    zusaetzliche g1_starts_geraet2-Alternierungs-Flag diente nur der
    Farbwahl und ist seit dem Gruppenzeiten-Umbau entfallen (siehe
    farbe_fuer_phase)."""
    plan_data = state.get("plan_data", {})
    entfall   = set(state.get("entfall_published", []))
    fixed_entries = fixed_entries or {}
    valid = []
    for datum_kurz, pdata in plan_data.items():
        if exclude_date and datum_kurz == exclude_date:
            continue
        if datum_kurz in entfall or not pdata:
            continue
        _fe = fixed_entries.get(datum_kurz) or {}
        g1 = (_fe.get("geraet_1") if isinstance(_fe, dict) else None) or pdata.get("geraet_1")
        g2 = (_fe.get("geraet_2") if isinstance(_fe, dict) else None) or pdata.get("geraet_2")
        idx = None
        for i, (r1, r2) in enumerate(GERAETE_ROTATION):
            if r1 == g1 and r2 == g2:
                idx = i; break
        if idx is None:
            continue
        try:
            d = datetime.strptime(datum_kurz, "%d.%m.%y").date()
        except Exception:
            continue
        valid.append((d, datum_kurz, idx))

    if not valid:
        print("[ROTATION] Kein gueltiger Geraete-Verlauf in plan_data -> Default Sprung+Reck.")
        return 1   # Sprung+Reck als sinnvoller Startpunkt

    valid.sort(key=lambda x: x[0])
    last8 = valid[-8:]
    _, last_key, last_idx = last8[-1]
    next_idx = (last_idx + 1) % len(GERAETE_ROTATION)
    verlauf = ", ".join(f"{k}={GERAETE_ROTATION[i][0]}+{GERAETE_ROTATION[i][1]}"
                        for _, k, i in last8)
    print(f"[ROTATION] Letzte {len(last8)} echten Trainings: {verlauf}")
    print(f"[ROTATION] Juengster echter Plan {last_key} = "
          f"{GERAETE_ROTATION[last_idx][0]}+{GERAETE_ROTATION[last_idx][1]} "
          f"-> naechster: {GERAETE_ROTATION[next_idx][0]}+{GERAETE_ROTATION[next_idx][1]}")
    return next_idx

def active_training_date():
    """Aktuell relevanter Trainingstag.
    - Trainingstag (Mi/Fr) TAGSUEBER: HEUTE (damit ein am Trainingstag markierter
      Entfall noch veroeffentlicht wird und Updates auf den heutigen Plan wirken).
    - Trainingstag ABENDS im Veroeffentlichungsfenster (>=22:00 CEST, Training
      vorbei): der NAECHSTE Trainingstag -> Mi-Abend erzeugt bereits den
      Freitagsplan, Fr-Abend den naechsten Mittwochsplan.
    - Sonst: der naechste Mi/Fr."""
    t = date.today()
    if t.weekday() in (2, 4) and not is_publication_window():
        return t
    return next_training_date()


def next_training_date():
    """Nächster Trainingstag (Mi/Fr) ab morgen."""
    d = date.today() + timedelta(days=1)
    while d.weekday() not in (2, 4):
        d += timedelta(days=1)
    return d

def fmt_datum(d):
    return d.strftime("%d.%m.%Y")

def fmt_datum_kurz(d):
    return d.strftime("%d.%m.%y")

def wochentag_name(d):
    return "Mittwoch" if d.weekday() == 2 else "Freitag"

# ════════════════════════════════════════════════════════════════
#  SFTP
# ════════════════════════════════════════════════════════════════

def get_sftp():
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(SSH_HOST, port=SSH_PORT, username=SSH_USER,
                   password=SSH_PASSWORD, timeout=20)
    return client, client.open_sftp()

def plan_exists(sftp, datum_kurz):
    try:
        sftp.stat(f"trainingspläne/{datum_kurz}_Trainingsplan.pdf")
        return True
    except FileNotFoundError:
        return False

def normalize_name(name):
    """Website-Format → interner Anzeigename (z.B. 'Vorname (G1)' → 'Vorname X.')."""
    # Schritt 1: "(G1)" Suffix entfernen → "Vorname G1"
    intermediate = re.sub(r'\s*\(G(\d+)\)$', r' G\1', name.strip())
    # Schritt 2: Mapping auf neues Anzeigeformat
    return WEBSITE_TO_DISPLAY.get(intermediate, WEBSITE_TO_DISPLAY.get(name.strip(), intermediate))

def read_abmeldungen(sftp):
    f    = sftp.open("abmeldungen/abmeldungen.json", "r")
    raw  = f.read()
    f.close()
    data = json.loads(raw.decode("utf-8", errors="replace"))
    raw_hash = hashlib.md5(raw).hexdigest()
    return data, raw_hash

def read_anmerkungen_server(sftp):
    """Liest ungelesene Trainer-Anmerkungen vom Server."""
    try:
        f    = sftp.open("anmerkungen/anmerkungen.json", "r")
        data = json.loads(f.read().decode("utf-8", errors="replace"))
        f.close()
        return [a for a in data if not a.get("gelesen", False)]
    except Exception:
        return []

def read_fixed_entries(sftp):
    """Liest fixed_entries.json vom Server (falls vorhanden).
    Struktur: { "DD.MM.JJ": { "lock_trainer_plan": bool,
                               "fixed_absences": { "Gruppe": ["Name", ...] },
                               "fixed_trainer_plan": { "Trainer": [[text,farbe],...] } } }
    """
    try:
        f    = sftp.open("fixed_entries.json", "r")
        data = json.loads(f.read().decode("utf-8", errors="replace"))
        f.close()
        print(f"[FIXED] fixed_entries.json geladen: {list(data.keys())}")
        return data
    except Exception:
        return {}

def apply_fixed_absences(absences, fixed_entry):
    """Merged fixed_absences in die berechneten Abwesenheiten (addiert, entfernt nichts)."""
    for gruppe, namen in fixed_entry.get("fixed_absences", {}).items():
        existing = absences.get(gruppe, [])
        for name in namen:
            if name not in existing:
                existing.append(name)
                print(f"[FIXED] {name} ({gruppe}) als fix-abwesend hinzugefügt.")
        absences[gruppe] = existing
    return absences

# ════════════════════════════════════════════════════════════════
#  TRAININGSENTFALL (Training abgesagt – von Trainer auf der Website markiert)
# ════════════════════════════════════════════════════════════════

def read_trainingsentfall(sftp):
    """Liest /abmeldungen/trainingsentfall.json (Liste von Y-m-d Strings).
    Single Source of Truth für abgesagte Trainings – wird auf der Website gepflegt."""
    try:
        f    = sftp.open("abmeldungen/trainingsentfall.json", "r")
        data = json.loads(f.read().decode("utf-8", errors="replace"))
        f.close()
        if isinstance(data, list):
            dates = [d for d in data if isinstance(d, str) and re.match(r"^\d{4}-\d{2}-\d{2}$", d)]
            print(f"[ENTFALL] trainingsentfall.json geladen: {dates}")
            return dates
    except Exception:
        pass
    return []

def build_entfall_pdf(datum, datum_kurz, wochentag, notiz=""):
    """Erzeugt den minimalen Trainingsentfall-Hinweis (xlsx+pdf), identisch zum
    manuellen 12.06-Hinweis. Gibt (xlsx_path, pdf_path) zurück."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trainingsplan"
    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 13.0
    ws.merge_cells("B1:H1"); ws.merge_cells("B2:H2"); ws.merge_cells("B3:H3")
    c = ws["B1"]; c.value = f"TRAININGSPLAN | {wochentag}, {datum}"
    c.fill = fill("2C3E50"); c.font = font(bold=True, color="FFFFFF", size=14); c.alignment = align()
    ws.row_dimensions[1].height = 30
    c = ws["B2"]; c.value = "⚠  TRAININGSENTFALL"
    c.fill = fill("C0392B"); c.font = font(bold=True, color="FFFFFF", size=22); c.alignment = align()
    ws.row_dimensions[2].height = 50
    notiz = (notiz or "").strip()
    b3_text = notiz if notiz else f"Das Training am {datum} ist ausgefallen."
    c = ws["B3"]; c.value = b3_text
    c.fill = fill("FAD7A0"); c.font = font(bold=False, color="000000", size=12); c.alignment = align(wrap=bool(notiz))
    # Zeilenhoehe an Anmerkungs-Laenge anpassen (mehrzeilige Anmerkungen)
    _lines = b3_text.count("\n") + 1 + (len(b3_text) // 55)
    ws.row_dimensions[3].height = max(28, _lines * 20)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:I4"

    out_dir = "/tmp/trainingsplan"; os.makedirs(out_dir, exist_ok=True)
    xlsx_path = os.path.join(out_dir, f"{datum_kurz}_Trainingsplan.xlsx")
    pdf_path  = os.path.join(out_dir, f"{datum_kurz}_Trainingsplan.pdf")
    wb.save(xlsx_path)
    tmp_dir = tempfile.mkdtemp()
    result  = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", tmp_dir, xlsx_path],
        capture_output=True, text=True, timeout=90)
    tmp_pdf = os.path.join(tmp_dir, os.path.basename(xlsx_path).replace(".xlsx", ".pdf"))
    if result.returncode == 0 and os.path.exists(tmp_pdf):
        shutil.copy2(tmp_pdf, pdf_path)
    else:
        raise RuntimeError(f"LibreOffice Fehler (Entfall): {result.stderr[:300]}")
    return xlsx_path, pdf_path

def publish_entfall(sftp, datum, datum_kurz, wochentag, notiz=""):
    """Veröffentlicht den Entfall-Hinweis: ersetzt einen evtl. vorhandenen normalen
    Plan durch den Trainingsentfall-Hinweis und aktualisiert das Widget-JSON.
    notiz = im Admin-Bereich (fixed_entries) eingetragene Anmerkung -> wird Inhalt."""
    xlsx_path, pdf_path = build_entfall_pdf(datum, datum_kurz, wochentag, notiz)
    upload_pdf(sftp, pdf_path, datum_kurz)
    upload_xlsx(sftp, xlsx_path, datum_kurz)
    tag, monat, jahr = datum.split(".")
    datum_iso = f"{jahr}-{monat}-{tag}"
    aktuell = {
        "datum": datum, "datum_iso": datum_iso, "wochentag": wochentag,
        "trainingsentfall": True,
        "pdf_url": f"https://tv-rheinzabern.e-websolutions.de/trainingspläne/{datum_kurz}_Trainingsplan.pdf",
        "einteilung": {},
    }
    upload_aktuell_json(sftp, aktuell)
    print(f"[ENTFALL] Entfall-Hinweis für {datum} veröffentlicht.")

def remove_plan_files(sftp, datum_kurz):
    """Entfernt PDF/XLSX eines Plans vom Server (z.B. wenn Entfall aufgehoben wird,
    damit ein frischer normaler Plan erzeugt wird)."""
    for ext in ("pdf", "xlsx"):
        try:
            sftp.remove(f"trainingspläne/{datum_kurz}_Trainingsplan.{ext}")
            print(f"[ENTFALL] Entfernt: trainingspläne/{datum_kurz}_Trainingsplan.{ext}")
        except Exception:
            pass

def _entfall_is_recent(iso_date, today, max_age_days=14):
    """Wie alt ist ein trainingsentfall.json-Eintrag, gemessen an `today`?
    True nur fuer Eintraege innerhalb der letzten `max_age_days` Tage (auch
    kuenftige Daten zaehlen als 'recent', negative Alterswerte inklusive).
    Filtert die Monate alten Karteileichen aus, die trainingsentfall.json nie
    entfernt (siehe Hotfix-Kommentar in main())."""
    try:
        d = datetime.strptime(iso_date, "%Y-%m-%d").date()
    except Exception:
        return False
    return (today - d).days <= max_age_days

def _publish_entfall_for(sftp, state, fixed_entries, iso_date, entfall_published):
    """Veroeffentlicht (falls noetig) den Entfall-Hinweis fuer EIN einzelnes Datum aus
    trainingsentfall.json. Ausgelagert aus main(), weil der Entfall-Check bisher nur
    das naechste anstehende Training (active_training_date()) prueft -- jeder andere
    Eintrag in trainingsentfall.json (z.B. ein bereits vergangenes, verspaetet
    markiertes Training) wurde dadurch nie veroeffentlicht, siehe main(). Gibt True
    zurueck, wenn tatsaechlich (neu) veroeffentlicht wurde."""
    try:
        d = datetime.strptime(iso_date, "%Y-%m-%d").date()
    except Exception:
        return False
    dk    = fmt_datum_kurz(d)
    datum = fmt_datum(d)
    wtag  = wochentag_name(d)
    fixed = fixed_entries.get(dk, {})
    notiz = (fixed.get("notiz", "") or "").strip()
    nhash = hashlib.md5(notiz.encode("utf-8")).hexdigest()
    notiz_store   = state.setdefault("entfall_notiz_hash", {})
    notiz_changed = notiz_store.get(dk) != nhash
    was_published = dk in entfall_published
    if was_published and plan_exists(sftp, dk) and not notiz_changed:
        return False
    publish_entfall(sftp, datum, dk, wtag, notiz)
    notiz_store[dk] = nhash
    if dk not in entfall_published:
        entfall_published.append(dk)
    if dk not in state.setdefault("generated_plans", []):
        state["generated_plans"].append(dk)
    state.get("plan_data", {}).pop(dk, None)  # alten Plan-Hash verwerfen
    if not was_published:
        send_whatsapp(
            f"Hi, Cloude hier ⚠️\n\n"
            f"Das Training am {wtag}, {datum} ist als Trainingsentfall markiert.\n"
            f"Ich habe den Entfall-Hinweis veröffentlicht und erstelle KEINEN normalen Plan."
        )
    print(f"[ENTFALL] Entfall für {datum} verarbeitet.")
    return True

def mark_anmerkungen_gelesen(sftp, ids_to_mark):
    """Markiert angegebene Anmerkungen als gelesen auf dem Server."""
    if not ids_to_mark:
        return
    try:
        f    = sftp.open("anmerkungen/anmerkungen.json", "r")
        data = json.loads(f.read().decode("utf-8", errors="replace"))
        f.close()
        for entry in data:
            if entry.get("id") in ids_to_mark:
                entry["gelesen"] = True
        f = sftp.open("anmerkungen/anmerkungen.json", "w")
        f.write(json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8"))
        f.close()
        print(f"[OK] {len(ids_to_mark)} Anmerkung(en) als gelesen markiert.")
    except Exception as e:
        print(f"[WARN] Anmerkungen konnten nicht markiert werden: {e}")

def upload_pdf(sftp, local_path, datum_kurz):
    remote = f"trainingspläne/{datum_kurz}_Trainingsplan.pdf"
    sftp.put(local_path, remote)
    print(f"[OK] Hochgeladen: {remote}")

def upload_xlsx(sftp, local_path, datum_kurz):
    remote = f"trainingspläne/{datum_kurz}_Trainingsplan.xlsx"
    sftp.put(local_path, remote)
    print(f"[OK] Hochgeladen: {remote}")

def build_aktuell_json(datum, datum_kurz, wochentag, geraet_1, geraet_2, trainer_plan, abwesend, grid_rows):
    """Erstellt trainingsplan_aktuell.json für iOS Widgets."""
    tag, monat, jahr = datum.split(".")
    datum_iso = f"{jahr}-{monat}-{tag}"
    pdf_url = f"https://tv-rheinzabern.e-websolutions.de/trainingspläne/{datum_kurz}_Trainingsplan.pdf"
    einteilung = {}
    for trainer, plan in trainer_plan.items():
        if plan is None:
            continue
        if trainer in abwesend.get("Trainer", []):
            continue
        slots = []
        for slot_idx, (row_start, _row_end) in enumerate(grid_rows):
            if slot_idx < len(plan):
                text, _ = plan[slot_idx]
                slots.append({"zeit_start": _min_to_hhmm(row_start), "aufgabe": text})
        einteilung[trainer] = slots
    return {
        "datum": datum, "datum_iso": datum_iso, "wochentag": wochentag,
        "geraet1": geraet_1, "geraet2": geraet_2, "pdf_url": pdf_url,
        "einteilung": einteilung,
    }

def upload_aktuell_json(sftp, json_data):
    """Lädt trainingsplan_aktuell.json auf Server hoch (für iOS Widgets)."""
    data = json.dumps(json_data, indent=2, ensure_ascii=False).encode("utf-8")
    f = sftp.open("trainingspläne/trainingsplan_aktuell.json", "wb")
    f.write(data)
    f.close()
    print("[OK] Hochgeladen: trainingspläne/trainingsplan_aktuell.json")

# ════════════════════════════════════════════════════════════════
#  DYNAMISCHES ZEITRASTER (ersetzt das frühere feste 5-Slot-Raster)
# ════════════════════════════════════════════════════════════════

def _hhmm_to_min(s):
    h, m = str(s).split(":")
    return int(h) * 60 + int(m)

def _min_to_hhmm(m):
    return f"{m // 60:02d}:{m % 60:02d}"

def tag_of_date(d):
    """'mi'/'fr' aus einem Trainingsdatum -- Schluessel in GRUPPEN_ZEITEN."""
    return "mi" if d.weekday() == 2 else "fr"

def compute_time_grid(gruppen_zeiten, tag):
    """Baut das dynamische Zeitraster fuer einen Wochentag ('mi'/'fr') aus den
    konfigurierten Gruppenzeiten: sammelt alle Start-/End-Minuten aller
    Gruppen-Phasen, bildet daraus die sortierte, lueckenlose Zeilenfolge.
    Ersetzt das fruehere feste 5-Zeilen-Raster (SLOT_START_MIN/SLOT_END_MIN).
    Fuer die migrierten Default-Zeiten (G1/G2/G3 17:00-17:30/17:30-18:15/
    18:15-19:00, G4 versetzt) liefert das 4 statt der frueheren 5 Zeilen --
    die alte fuenfte Zeile (18:00-18:15) war eine willkuerliche Unterteilung
    MITTEN in der durchgehenden Geraet1-Phase (17:30-18:15) ohne eigene
    Bedeutung (dieselbe Farbe/derselbe Text in beiden Teil-Zeilen). Das neue
    Raster leitet Zeilen ausschliesslich aus echten Phasengrenzen ab, wodurch
    dieser doppelte Eintrag entfaellt -- inhaltlich (Farben/Zeiten/Text)
    aendert sich dadurch nichts, der gedruckte Plan wird nur um eine
    redundante Zeile kuerzer.

    Gibt (rows, phase_by_gruppe) zurueck:
      rows            = [(start_min, end_min), ...] sortiert, lueckenlos
      phase_by_gruppe = {gruppe: ["aufwaermen"|"geraet1"|"geraet2"|None, ...]} je Zeile
    """
    boundaries = set()
    for tage in (gruppen_zeiten or {}).values():
        zeiten = tage.get(tag) or {}
        for phase in ("aufwaermen", "geraet1", "geraet2"):
            fenster = zeiten.get(phase)
            if not fenster:
                continue
            boundaries.add(_hhmm_to_min(fenster["start"]))
            boundaries.add(_hhmm_to_min(fenster["ende"]))
    bounds = sorted(boundaries)
    rows = list(zip(bounds, bounds[1:]))
    phase_by_gruppe = {}
    for g, tage in (gruppen_zeiten or {}).items():
        zeiten = tage.get(tag) or {}
        row_phases = []
        for (s, e) in rows:
            mid = (s + e) // 2
            phase = None
            for ph in ("aufwaermen", "geraet1", "geraet2"):
                fenster = zeiten.get(ph)
                if fenster and _hhmm_to_min(fenster["start"]) <= mid < _hhmm_to_min(fenster["ende"]):
                    phase = ph
                    break
            row_phases.append(phase)
        phase_by_gruppe[g] = row_phases
    return rows, phase_by_gruppe

def group_active_rows(row_phases):
    """Erster/letzter Zeilen-Index einer Gruppe mit einer echten Phase (nicht
    None) -- fuer die Aufbauen/Abbauen-Randzeilen (z.B. wartet eine Gruppe,
    die spaeter beginnt, die erste Zeile mit Aufbauen statt Aufwaermen)."""
    active = [i for i, p in enumerate(row_phases) if p is not None]
    if not active:
        return None, None
    return active[0], active[-1]

def zeiten_kompatibel(gruppen_zeiten, a, b):
    """True, wenn zwei Gruppen an BEIDEN Wochentagen exakt dieselben Phasen-
    Zeitfenster UND dieselbe Geraete-Reihenfolge (GRUPPEN_TAUSCH) haben --
    Voraussetzung fuer eine Zusammenlegung (ein Trainer kann nicht
    gleichzeitig zwei verschiedene Zeitplaene halten, und eine zusammengelegte
    Einheit turnt als EIN Trainer-gefuehrter Block auf einem eindeutigen
    Geraet -- bei unterschiedlichem Tausch waere unklar, welches)."""
    if (a in GRUPPEN_TAUSCH) != (b in GRUPPEN_TAUSCH):
        return False
    za, zb = gruppen_zeiten.get(a) or {}, gruppen_zeiten.get(b) or {}
    for tag in ("mi", "fr"):
        ta, tb = za.get(tag) or {}, zb.get(tag) or {}
        for phase in ("aufwaermen", "geraet1", "geraet2"):
            if (ta.get(phase) or {}) != (tb.get(phase) or {}):
                return False
    return True

# ════════════════════════════════════════════════════════════════
#  ABWESENHEITEN AUSWERTEN
# ════════════════════════════════════════════════════════════════

def _extract_time_min(notiz):
    """Findet eine Uhrzeit (HH:MM, HH.MM, 'HH Uhr') -> Minuten seit 0:00 oder None."""
    s = (notiz or "").lower()
    m = re.search(r'(\d{1,2})[:.](\d{2})', s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if 15 <= h <= 21 and 0 <= mi < 60:
            return h*60+mi
    m = re.search(r'(\d{1,2})\s*uhr', s)
    if m:
        h = int(m.group(1))
        if 15 <= h <= 21:
            return h*60
    return None

_RE_MUSS_FRUEH = re.compile(r'\bmuss\b.{0,25}?\b(?:los|weg|gehen)\b')

def parse_trainer_timing(notiz):
    """Richtung ('spaet'=kommt spaeter / 'frueh'=geht frueher / None) + Uhrzeit-Min."""
    s = (notiz or "").lower()
    t = _extract_time_min(s)
    frueh_kw = ["früher", "frueher", "geht", "gehe", "gehen", "weg", "raus", "verlass",
                "eher", "nur bis", "muss los", "muss weg", "muss gehen", "vorzeitig"]
    spaet_kw = ["später", "spaeter", "verspät", "verspaet", "komme", "kommt", "erst",
                "etwas spät", "bisschen spät", "verspätung", "verzögert"]
    is_frueh = any(k in s for k in frueh_kw) or bool(_RE_MUSS_FRUEH.search(s))
    is_spaet = any(k in s for k in spaet_kw)
    if "bis" in s and t is not None:
        return "frueh", t
    if is_frueh and not is_spaet:
        return "frueh", t
    if is_spaet and not is_frueh:
        return "spaet", t
    if is_spaet and is_frueh:
        if "komme" in s or "kommt" in s or "erst" in s:
            return "spaet", t
        return "frueh", t
    return None, t

def is_timing_note(notiz, kind, tmin):
    """True = Verspaetung/frueher-gehen (Trainer bleibt da). Echte Abwesenheiten
    ('Urlaub','krank','Spätschicht') liefern kind=None -> False."""
    if kind is None:
        return False
    if tmin is not None:
        return True
    s = (notiz or "").lower()
    strong = ["später", "spaeter", "früher", "frueher", "verspät", "verspaet",
              "kommt erst", "komme erst", "geht früh", "geh früh", "nur bis"]
    return any(k in s for k in strong)

def compute_blocked_slots(kind, time_min, grid_rows):
    """Set geblockter Zeilen-Indizes im dynamischen Raster. spaet: bis Ankunft;
    frueh: ab Weggang."""
    n = len(grid_rows)
    if not n:
        return set()
    if kind == "spaet":
        if time_min is None:
            return {0}
        return {i for i, (s, _e) in enumerate(grid_rows) if time_min > s}
    if kind == "frueh":
        if time_min is None:
            return {n - 1}
        return {i for i, (_s, e) in enumerate(grid_rows) if time_min < e}
    return set()

def upload_anmerkungen_auto(sftp, datum_kurz, lines):
    """Schreibt die generierten Auto-Anmerkungen je Datum nach anmerkungen_auto.json
    (Quelle fuer die Vorbefuellung des Notizfeldes in admin.php)."""
    try:
        try:
            f = sftp.open("anmerkungen_auto.json", "r")
            data = json.loads(f.read().decode("utf-8"))
            f.close()
            if not isinstance(data, dict):
                data = {}
        except Exception:
            data = {}
        data[datum_kurz] = list(lines)
        f = sftp.open("anmerkungen_auto.json", "wb")
        f.write(json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8"))
        f.close()
        print(f"[ANM] anmerkungen_auto.json aktualisiert ({datum_kurz}: {len(lines)} Zeilen).")
    except Exception as e:
        print(f"[ANM] anmerkungen_auto.json nicht geschrieben: {e}")

def timing_annotations(trainer_timing):
    """Anmerkungs-Zeilen: '(Name) kommt später (HH:MM): Notiz'."""
    lines = []
    for name, info in (trainer_timing or {}).items():
        label = "kommt später" if info.get("kind") == "spaet" else "geht früher"
        ts    = info.get("time_str")
        notiz = (info.get("notiz") or "").strip()
        head  = f"{name} {label}" + (f" ({ts})" if ts else "")
        lines.append(f"• {head}: {notiz}" if notiz else f"• {head}")
    return lines

def _find_free_coverer(trainer_plan, trainer_timing, slot, exclude):
    """Freien Trainer fuer 'slot' finden: erst Springer, dann Auf-/Abbau.
    Wer in diesem Slot selbst spaet/frueh geblockt ist, scheidet aus."""
    def blocked_at(t):
        info = trainer_timing.get(t)
        return bool(info) and slot in info.get("blocked", [])
    for want in ("springer", "aufbauen"):
        for t, plan in trainer_plan.items():
            if t == exclude or not plan or slot >= len(plan):
                continue
            if blocked_at(t):
                continue
            _txt, ck = plan[slot]
            if ck == want:
                return t
    return None

def apply_timing_coverage(trainer_plan, trainer_timing):
    """Vertretung: geht ein Trainer frueher / kommt spaeter, uebernimmt fuer die
    fehlenden Slots ein freier Trainer (bevorzugt der Springer) dessen Gruppe.
    MUSS VOR apply_timing_blocks laufen (liest die Original-Gruppenzelle)."""
    if not trainer_timing:
        return trainer_plan
    GRUPPEN_COLORS = {"aufwaermen", "g1_blau", "g2_orange"}
    for name, info in trainer_timing.items():
        plan = trainer_plan.get(name)
        if not plan:
            continue
        for i in sorted(info.get("blocked", [])):
            if not (0 <= i < len(plan)):
                continue
            text, ck = plan[i]
            if ck not in GRUPPEN_COLORS:
                continue
            coverer = _find_free_coverer(trainer_plan, trainer_timing, i, exclude=name)
            if coverer:
                p = list(trainer_plan[coverer]); p[i] = (text, ck); trainer_plan[coverer] = p
    return trainer_plan


def apply_timing_blocks(trainer_plan, trainer_timing):
    """Ueberschreibt geblockte Slots verspaeteter/frueher gehender Trainer ROT mit
    'kommt später'/'geht früher'. Trainer bleibt eingeteilt (NICHT abwesend)."""
    if not trainer_timing:
        return trainer_plan
    for name, info in trainer_timing.items():
        plan = trainer_plan.get(name)
        if not plan:
            continue
        label = "kommt später" if info.get("kind") == "spaet" else "geht früher"
        ts    = info.get("time_str")
        cell  = f"{label}\n{ts}" if ts else label
        new   = list(plan)
        for i in info.get("blocked", []):
            if 0 <= i < len(new):
                new[i] = (cell, "sonder")
        trainer_plan[name] = new
    return trainer_plan

def get_absences(abmeldungen, training_date, grid_rows):
    target    = training_date.strftime("%Y-%m-%d")
    absences  = {g: [] for g in GRUPPEN_ORDER}
    absences["Trainer"] = []
    late_notes = []
    trainer_timing = {}

    for entry in abmeldungen:
        if entry.get("datum") != target:
            continue
        name   = normalize_name(entry.get("name", "").strip())
        gruppe = entry.get("gruppe", "").strip()
        notiz  = entry.get("notiz", "").strip()

        # Trainer mit Verspaetung / frueher-gehen: ANWESEND, nur Zeitbloecke blocken.
        # Die Website hat zwei eigene Kontrollkaestchen ("Verspaetung"/"Frueher gehen"),
        # die unabhaengig vom Notiz-Text gesetzt werden und Vorrang vor der reinen
        # Text-Erkennung haben (Bugfix 28.08.2026: frueher_gehen wurde bisher nirgends
        # gelesen, wodurch ein Trainer mit gesetztem Haken aber unpassendem/fehlendem
        # Notiz-Text faelschlich als komplett abwesend statt als Teil-Anwesend galt).
        if name in ALLE_TRAINER:
            kind, tmin = parse_trainer_timing(notiz)
            is_versp      = bool(entry.get("verspaetung", False))
            is_frueh_flag = bool(entry.get("frueher_gehen", False))
            if is_versp or is_frueh_flag or is_timing_note(notiz, kind, tmin):
                if is_versp:
                    kind = "spaet"
                elif is_frueh_flag:
                    kind = "frueh"
                elif kind is None:
                    kind = "spaet"
                blocked = compute_blocked_slots(kind, tmin, grid_rows)
                tstr = f"{tmin//60:02d}:{tmin%60:02d}" if tmin is not None else None
                trainer_timing[name] = {
                    "kind": kind, "time_min": tmin, "time_str": tstr,
                    "notiz": notiz, "blocked": sorted(blocked),
                }
                continue

        if gruppe in absences:
            absences[gruppe].append(name)

    return absences, late_notes, trainer_timing

# ════════════════════════════════════════════════════════════════
#  KOMPLEXE FÄLLE ERKENNEN
# ════════════════════════════════════════════════════════════════

def detect_complex(absences, late_notes):
    # Es wird IMMER ein Plan erstellt. Staffing-Probleme sind nur soft_warnings
    # (Hinweis per WhatsApp/Mail) -> autonomer Plan auch bei wenig Trainern/Kindern.
    issues        = []
    soft_warnings = []
    anwesend_trainer = [t for t in ALLE_TRAINER if t not in absences.get("Trainer", [])]
    n = len(anwesend_trainer)
    if n <= 3:
        soft_warnings.append((f"low_trainer_{n}",
            f"Nur {n} Trainer anwesend ({', '.join(anwesend_trainer) or '-'}). "
            f"Gruppen werden automatisch sinnvoll zusammengelegt."))
    bekannte_namen = set(ALLE_TRAINER)
    for turner in ALLE_TURNER.values():
        bekannte_namen.update(turner)
    for gruppe, names in absences.items():
        for name in names:
            if name not in bekannte_namen:
                soft_warnings.append((f"unknown_{gruppe}_{name}",
                    f"Unbekannter Name: '{name}' (Gruppe {gruppe})"))
    for gruppe in GRUPPEN_ORDER:
        turner = ALLE_TURNER.get(gruppe, [])
        abw = absences.get(gruppe, [])
        anwesend = len(turner) - len(abw)
        if len(abw) >= len(turner):
            soft_warnings.append((f"empty_{gruppe}", f"Gruppe {gruppe}: alle Turner abwesend."))
        elif anwesend <= 2:
            soft_warnings.append((f"low_turner_{gruppe}_{anwesend}",
                f"Gruppe {gruppe} hat nur {anwesend} Turner anwesend."))
    return issues, anwesend_trainer, soft_warnings

# ════════════════════════════════════════════════════════════════
#  GERAETE-FARBEN / ZELLEN: EINZIGE Quelle fuer ALLE Trainer-Plan-Builder
#  (Standard-Pfad build_trainer_plan, KI-Pfad build_ki_einteilung, Not-
#  besetzung _build_lowstaff_plan, Admin-Fallback build_admin_trainer_plan).
#  Bugfix 12./19.08.2026 (Noah): vorher hatte jeder Builder seine eigene
#  Kopie dieser Zuordnung und lief bei bestimmten Konstellationen auseinander
#  (Farb-Kollisionen zwischen Gruppen). Seit dem Gruppenzeiten-Umbau ist das
#  strukturell ausgeschlossen: nur noch 2 Farben (Geraet1/Geraet2), und eine
#  Zusammenlegung ist ueberhaupt nur zwischen zeit-kompatiblen Gruppen
#  erlaubt (siehe zeiten_kompatibel), wodurch nie zwei Einheiten dieselbe
#  Zeile gleichzeitig beanspruchen koennen.
# ════════════════════════════════════════════════════════════════

def _cells_for_unit(label, groups, grid_rows, grid_phase):
    """Baut die Zellen-Liste (Text, Farbschluessel) je Zeitraster-Zeile fuer
    eine Trainer-Einheit (eine Gruppe oder mehrere zeit-kompatible Gruppen
    zusammengelegt -- alle Gruppen der Einheit haben laut zeiten_kompatibel
    exakt dasselbe Phasenmuster, jede von ihnen kann daher als Referenz
    dienen)."""
    ref = groups[0]
    ref_phases = grid_phase.get(ref) or [None] * len(grid_rows)
    first_active, last_active = group_active_rows(ref_phases)
    cells = []
    for i in range(len(grid_rows)):
        phase = ref_phases[i] if i < len(ref_phases) else None
        if phase == "aufwaermen":
            cells.append((f"AW {label}", "aufwaermen"))
        elif phase in ("geraet1", "geraet2"):
            cells.append((label, farbe_fuer_phase(_effektive_phase(ref, phase))))
        elif first_active is not None and i > last_active:
            cells.append(("Abbauen", "aufbauen"))
        else:
            cells.append(("Aufbauen", "aufbauen"))
    return cells

def _springer_cells(grid_rows):
    n = len(grid_rows)
    if n == 0:
        return []
    if n == 1:
        return [("Springer", "springer")]
    return [("Aufbauen", "aufbauen")] + [("Springer", "springer")] * (n - 2) + [("Abbauen", "aufbauen")]

def unit_label(groups):
    return "+".join(groups)

# ────────────────────────────────────────────────────────────────
#  ROLLEN-ROTATION -- echte Fairness ueber ALLE Trainer und ALLE Rollen
# ────────────────────────────────────────────────────────────────
# Ersetzt die fruehere Teil-Rotation (nur im Rest-Pool nach fest verdrahteten
# Trainer-Praeferenzen) durch eine gemeinsame Fairness-Funktion, die von
# BEIDEM Pfaden (Standard build_trainer_plan UND KI-Pfad build_ki_einteilung)
# genutzt wird -- bewusst eine gemeinsame Funktion statt zwei Kopien, da
# genau diese Art Duplizierung in diesem Projekt bereits mehrfach zu
# Drift-Bugs gefuehrt hat (siehe Farb-Kollisions-Historie).
# Speicherung weiterhin ueber plan_data[datum_kurz]["trainer_roles"] =
# {Trainer: "Springer" | <Einheiten-Label z.B. "G1" oder "G2+G3"> | None},
# aber ueber die letzten (bis zu) 6 Termine aggregiert statt nur den letzten
# (siehe _load_trainer_roles_history).

def _extract_trainer_roles(trainer_plan):
    """Ermittelt aus einem generierten trainer_plan-Dict die Rolle jedes
    Trainers: 'Springer' | <Gruppen-Einheiten-Label> | None. Das konkrete
    Label (nicht nur ein generisches 'Gruppe') wird gebraucht, damit die
    Rotations-Fairness weiss, WELCHE Einheit ein Trainer zuletzt hatte."""
    roles = {}
    for t in ALLE_TRAINER:
        cells = (trainer_plan or {}).get(t)
        if not cells:
            roles[t] = None
            continue
        labels = {str(c[0] or "").strip() for c in cells if isinstance(c, (list, tuple))}
        if labels and labels <= {"Aufbauen", "Abbauen", "Springer"}:
            roles[t] = "Springer"
            continue
        gruppen_label = None
        for c in cells:
            if not isinstance(c, (list, tuple)) or len(c) < 2:
                continue
            text, farbe = c[0], c[1]
            if farbe in ("g1_blau", "g2_orange") and text:
                gruppen_label = str(text).strip()
                break
        roles[t] = gruppen_label or "Gruppe"
    return roles

def _load_trainer_roles_history(state, exclude_date=None, limit=6):
    """Baut {trainer: [{'date':..., 'role':...}, ...]} (aelteste zuerst) aus
    den letzten `limit` plan_data-Eintraegen mit gespeicherten trainer_roles
    -- Grundlage fuer _assign_units_fair(). Kompatibel mit dem alten
    Ein-Termin-Snapshot-Format (jeder Eintrag war schon {trainer: rolle}
    PRO Datum, hier wird nur ueber mehrere Termine aggregiert)."""
    pd = (state or {}).get("plan_data", {}) or {}
    entries = []
    for dk, d in pd.items():
        if exclude_date and dk == exclude_date:
            continue
        if not isinstance(d, dict):
            continue
        roles = d.get("trainer_roles")
        if not roles:
            continue
        try:
            entries.append((datetime.strptime(dk, "%d.%m.%y"), dk, roles))
        except Exception:
            pass
    entries.sort(key=lambda x: x[0])
    history = {}
    for _d, dk, roles in entries[-limit:]:
        for trainer, role in roles.items():
            if role is None:
                continue
            history.setdefault(trainer, []).append({"date": dk, "role": role})
    return history

def _role_groups(role):
    """Zerlegt ein gespeichertes Rollen-Label ('G1', 'G3+G4', 'Springer', None)
    in die Menge der zugrunde liegenden Gruppennamen ('Springer'/None -> leer)."""
    if not role or role == "Springer":
        return set()
    return set(role.split("+"))

def _assign_units_fair(units, candidate_trainers, history, immer_springer):
    """Verteilt `units` (Liste von Einheiten-Labels, z.B. ['G1','G2+G3']) so
    fair wie moeglich auf `candidate_trainers` ueber ALLE Rollen (welche
    Einheit ODER Springer). immer_springer-Trainer werden nur als letzte
    Instanz gezogen, wenn sonst nicht genug Trainer fuer alle Einheiten da
    sind (dann bevorzugt, wer am laengsten keine Gruppe hatte). Gibt
    (assign: {trainer: unit}, springers: [trainer]) zurueck."""
    def recs(t):
        return history.get(t, [])
    def times_springer(t):
        return sum(1 for r in recs(t) if r.get("role") == "Springer")
    def last_index_for(t, matcher):
        rs = recs(t)
        for i in range(len(rs) - 1, -1, -1):
            if matcher(rs[i]):
                return i - len(rs)   # negativ; je laenger her, desto kleiner
        return -999                  # nie -> hoechste Prioritaet
    def last_index_for_unit(t, unit_label):
        # Bugfix 30.08.2026 (Noah: "fuehlt sich an, als haette Noah immer G1 und
        # Andy immer G3+G4"): verglich bisher das Rollen-Label exakt ("G3+G4" ==
        # "G3+G4"), matchte also NIE, wenn eine Zusammenlegung mal anders ausfiel
        # (z.B. nur "G3" statt "G3+G4"). Ohne Ueberschneidung mit der Historie blieb
        # der Score fuer praktisch jeden Trainer dauerhaft -999 (nie), der
        # nachfolgende sort() ist dann stabil und faellt auf die Config-Reihenfolge
        # der Trainer zurueck -- macht die "faire" Rotation fuer Merge-Faelle
        # effektiv wirkungslos und erzeugt genau das gemeldete statische Muster.
        # Jetzt zaehlt jede Ueberschneidung der beteiligten Gruppen als Treffer.
        groups = set(unit_label.split("+"))
        return last_index_for(t, lambda r: bool(_role_groups(r.get("role")) & groups))

    normal_pool = [t for t in candidate_trainers if t not in immer_springer]
    forced_pool = [t for t in candidate_trainers if t in immer_springer]
    n_units = len(units)
    if len(normal_pool) >= n_units:
        pool, overflow_springers = list(normal_pool), list(forced_pool)
    else:
        need = n_units - len(normal_pool)
        forced_sorted = sorted(forced_pool,
            key=lambda t: last_index_for(t, lambda r: r.get("role") != "Springer"))
        pulled = forced_sorted[:need]
        pool = normal_pool + pulled
        overflow_springers = [t for t in forced_pool if t not in pulled]

    pool_sorted = sorted(pool,
        key=lambda t: (-times_springer(t), last_index_for(t, lambda r: r.get("role") != "Springer")))
    getters = pool_sorted[:n_units]
    springers = [t for t in pool if t not in getters] + overflow_springers

    # Bugfix 30.08.2026 (Noah: 'fuehlt sich an, als haette Noah immer G1 und
    # Andy immer G3+G4'): Die fruehere Zuteilung sortierte ALLE (Trainer,Einheit)-
    # Paare nach Score und griff dann gierig zu -- bei mehreren gleich guten
    # Paaren (haeufig, z.B. wenn zwei Trainer noch nie eine bestimmte Einheit
    # hatten) haengt das Ergebnis rein von der zufaelligen Iterationsreihenfolge
    # ab, nicht vom tatsaechlich global fairsten Ergebnis (ein Trainer kann
    # dabei in eine schlechtere Einheit gedraengt werden, obwohl eine andere,
    # global bessere Verteilung moeglich gewesen waere). Bei realistisch
    # kleinen Trainer-/Einheiten-Zahlen (Turnverein, keine Grossorganisation)
    # ist eine vollstaendige Suche ueber alle moeglichen Zuordnungen (Permutationen)
    # problemlos schnell und liefert garantiert die global fairste Zuteilung
    # (minimale Summe der Einzel-Scores).
    assign = {}
    if getters:
        best_assign, best_score = None, None
        for combo in itertools.permutations(units, len(getters)):
            total = sum(last_index_for_unit(t, u) for t, u in zip(getters, combo))
            if best_score is None or total < best_score:
                best_score, best_assign = total, dict(zip(getters, combo))
        assign = best_assign or {}
    return assign, springers

# ════════════════════════════════════════════════════════════════
#  TRAINER-EINTEILUNG
# ════════════════════════════════════════════════════════════════

def _build_lowstaff_plan(available, grid_rows, grid_phase, history, immer_springer):
    """Notbesetzung (<=3 Trainer): konfigurierte Gruppen werden auf so wenige
    Einheiten wie noetig zusammengelegt (nur zeit-kompatible Nachbarn), dann
    ueber die gemeinsame Fairness-Funktion zugeteilt -- alle Trainer gleich
    behandelt (keine Trainer-Praeferenz mehr)."""
    def _compatible(a, b):
        return all(zeiten_kompatibel(GRUPPEN_ZEITEN, x, y) for x in a for y in b)

    units = [[g] for g in GRUPPEN_ORDER]
    while len(units) > max(1, len(available)):
        merge_i = next((i for i in range(len(units) - 1) if _compatible(units[i], units[i+1])), None)
        if merge_i is None:
            break
        units[merge_i] = units[merge_i] + units[merge_i + 1]
        del units[merge_i + 1]

    unit_groups = {unit_label(u): u for u in units}
    labels = list(unit_groups.keys())
    assign, springers = _assign_units_fair(labels, available, history, immer_springer)

    TRAINER_PLAN = {}
    for t in ALLE_TRAINER:
        if t in assign:
            lab = assign[t]
            TRAINER_PLAN[t] = _cells_for_unit(lab, unit_groups[lab], grid_rows, grid_phase)
        elif t in springers:
            TRAINER_PLAN[t] = _springer_cells(grid_rows)
        else:
            TRAINER_PLAN[t] = None
    anmerkungen = ["Wenig Trainer da: Gruppen wurden automatisch zusammengelegt."]
    if not available:
        anmerkungen.append("ACHTUNG: Kein Trainer anwesend - bitte dringend klaeren!")
    return TRAINER_PLAN, {}, anmerkungen


def build_trainer_plan(absences, grid_rows, grid_phase, trainer_roles_history=None):
    abwesend  = absences.get("Trainer", [])
    available = [t for t in ALLE_TRAINER if t not in abwesend]
    n = len(available)
    trainer_roles_history = trainer_roles_history or {}

    # anwesende Kinder je Gruppe
    present = {g: max(0, len(ALLE_TURNER.get(g, [])) - len(absences.get(g, []))) for g in GRUPPEN_ORDER}
    active  = [g for g in GRUPPEN_ORDER if present[g] >= 1]

    anmerkungen = []

    if not active:
        anmerkungen.insert(0, "ACHTUNG: Alle Turner abwesend - kein Training.")
        return {t: None for t in ALLE_TRAINER}, {}, anmerkungen
    if n == 0:
        anmerkungen.insert(0, "ACHTUNG: Kein Trainer anwesend - bitte dringend klaeren!")
        return {t: None for t in ALLE_TRAINER}, {}, anmerkungen

    # Bugfix 27.08.2026 (Noah: "Cassi geht ab 18:30, der Plan nimmt sie aber
    # komplett raus, soll aber moeglichst lange eine Gruppe uebernehmen"):
    # frueher gehende/spaeter kommende Trainer sind normale Halter-Kandidaten,
    # NICHT von vornherein ausgeschlossen -- apply_timing_coverage() (siehe
    # main()) uebernimmt danach automatisch die geblockten Randzeilen ihrer
    # Gruppe an einen freien Trainer, apply_timing_blocks() faerbt sie rot mit
    # "kommt spaeter"/"geht frueher". Vorher wurden partielle Trainer hier
    # praeventiv aus dem Halter-Pool entfernt, sobald irgendein Vollzeit-
    # Trainer da war -- das machte die Coverage-Mechanik faktisch wirkungslos
    # (ein Trainer, der nie eine Gruppe bekommt, braucht auch keine Vertretung)
    # und war zusaetzlich inkonsistent mit build_ki_einteilung(), das
    # trainer_timing schon vorher gar nicht ausgeschlossen hat.
    holders_pool = list(available)

    MIN_KIDS = 3   # jede Gruppe braucht >= 3 anwesende Turner, sonst zusammenlegen

    def _compatible(a, b):
        return all(zeiten_kompatibel(GRUPPEN_ZEITEN, x, y) for x in a for y in b)

    units = [[g] for g in active]

    # 1) Mindestgroesse: zu kleine Gruppe (<3) mit einem zeit-kompatiblen
    #    Nachbarn zusammenlegen. Kein kompatibler Nachbar da -> Gruppe bleibt
    #    einzeln (Notlage-Fallback statt eines unmoeglichen Merges).
    def _merge_undersized():
        for i, u in enumerate(units):
            if sum(present[g] for g in u) < MIN_KIDS and len(units) > 1:
                cands = [j for j in (i - 1, i + 1) if 0 <= j < len(units) and _compatible(units[i], units[j])]
                if not cands:
                    continue
                best = min(cands, key=lambda j: sum(present[g] for g in units[i]) + sum(present[g] for g in units[j]))
                lo, hi = sorted((i, best))
                units[lo] = units[lo] + units[hi]
                del units[hi]
                return True
        return False
    while _merge_undersized():
        pass

    # 2) Nicht mehr Einheiten als Vollzeit-Halter -> weiter zusammenlegen
    #    (kleinste zeit-kompatible Paarung zuerst). Gibt es keinen
    #    kompatiblen Merge-Partner mehr, bleiben ueberzaehlige Einheiten
    #    einzeln (Notlage) -- werden dann unten als unbesetzt markiert.
    #
    #    Bugfix 30.08.2026 (Noah: ein als "immer nur Springer" markierter Trainer
    #    bekam trotzdem regelmaessig eine Gruppe): Ziel war bisher len(holders_pool)
    #    -- ALLE verfuegbaren Trainer inkl. immer_springer. Das merged nur so weit
    #    wie fuer die reine Kopfzahl noetig, nicht so weit wie fuer die Kopfzahl OHNE
    #    immer_springer-Trainer noetig waere -- _assign_units_fair() musste dadurch
    #    haeufiger als eigentlich noetig auf den immer_springer-Trainer zurueckgreifen.
    #    Jetzt wird zuerst versucht, so weit zusammenzulegen, dass die normalen
    #    (nicht immer_springer) Trainer allein ausreichen; nur wenn dafuer kein
    #    kompatibler Merge-Partner mehr existiert, greift wie gehabt der
    #    Notlage-Fallback in _assign_units_fair() (immer_springer wird doch gezogen).
    _normal_holders = [t for t in holders_pool if t not in IMMER_SPRINGER]
    _merge_target = len(_normal_holders) if _normal_holders else len(holders_pool)
    while len(units) > max(1, _merge_target):
        best = None
        for i in range(len(units)):
            for j in range(i + 1, len(units)):
                if not _compatible(units[i], units[j]):
                    continue
                score = sum(present[g] for g in units[i]) + sum(present[g] for g in units[j])
                if best is None or score < best[0]:
                    best = (score, i, j)
        if best is None:
            break
        _, i, j = best
        units[i] = units[i] + units[j]
        del units[j]

    unit_groups = {unit_label(u): u for u in units}
    labels = list(unit_groups.keys())
    assign, springers = _assign_units_fair(labels, holders_pool, trainer_roles_history, IMMER_SPRINGER)

    unassigned_units = [lab for lab in labels if lab not in assign]
    if unassigned_units:
        anmerkungen.append("ACHTUNG: kein Trainer mehr fuer: " + ", ".join(unassigned_units))

    # partielle (frueh/spaet) Trainer -> immer Springer
    for t in available:
        if t not in assign and t not in springers:
            springers.append(t)

    TRAINER_PLAN = {}
    for t in ALLE_TRAINER:
        if t in assign:
            lab = assign[t]
            TRAINER_PLAN[t] = _cells_for_unit(lab, unit_groups[lab], grid_rows, grid_phase)
        elif t in springers:
            TRAINER_PLAN[t] = _springer_cells(grid_rows)
        else:
            TRAINER_PLAN[t] = None

    return TRAINER_PLAN, {}, anmerkungen

# ════════════════════════════════════════════════════════════════
#  EXCEL AUFBAUEN
# ════════════════════════════════════════════════════════════════

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="FFFFFF", size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value, hex_fill=None, fnt=None, aln=None):
    c = ws.cell(row=row, column=col, value=value)
    if hex_fill: c.fill = fill(hex_fill)
    if fnt:      c.font = fnt
    if aln:      c.alignment = aln
    return c

def merge_set(ws, row, cs, ce, value, hex_fill, fnt, aln=None):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=value)
    c.fill = fill(hex_fill)
    c.font = fnt
    c.alignment = aln or align()
    return c

def build_excel(datum, wochentag, geraet_1, geraet_2, abwesend,
                trainer_plan, sondertiming, anmerkungen, grid_rows):
    datum_kurz = datum[0:2] + "." + datum[3:5] + "." + datum[8:10]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trainingsplan"

    trainer_col = 3 + len(GRUPPEN_ORDER)   # Spalte "TRAINER" in der Anwesenheitstabelle
    end_col     = trainer_col + 1          # rechte Randspalte (wie frueher Spalte 8 bei 4 Gruppen)
    n_cols      = max(9, end_col + 1)      # mind. A-I wie frueher, sonst dynamisch breiter

    for ci in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    row = 1

    merge_set(ws, row, 2, end_col,
        f"TRAININGSPLAN | {wochentag}, {datum}",
        FARBEN["titel"], font(bold=True, size=13), align(h="left"))
    row += 1

    merge_set(ws, row, 2, end_col,
        f"Gerät 1 = {geraet_1} | Gerät 2 = {geraet_2}",
        FARBEN["header"], font(size=10), align(h="left"))
    row += 1

    merge_set(ws, row, 2, end_col, "ANWESENHEITEN",
        FARBEN["titel"], font(bold=True, size=10))
    row += 1

    for col_idx, label in enumerate(GRUPPEN_ORDER + ["TRAINER"], start=3):
        set_cell(ws, row, col_idx, label, FARBEN["header"], font(bold=True))
    row += 1

    max_len = max([len(v) for v in ALLE_TURNER.values()] + [len(ALLE_TRAINER)])
    alt = ["EBF5FB","FFFFFF"]

    for i in range(max_len):
        a = alt[i % 2]
        for col_bg in [2, end_col]:
            ws.cell(row=row, column=col_bg).fill = fill("FFFFFF")

        for ci, gruppe in enumerate(GRUPPEN_ORDER, start=3):
            tlist = ALLE_TURNER.get(gruppe, [])
            if i < len(tlist):
                name = tlist[i]
                ab   = name in abwesend.get(gruppe, [])
                set_cell(ws, row, ci,
                         f"{'✗' if ab else '✓'} {name}",
                         FARBEN["abwesend_turner"] if ab else FARBEN["anwesend"],
                         font(size=9, color="FFFFFF"), align(h="left"))
            else:
                ws.cell(row=row, column=ci).fill = fill(a)

        if i < len(ALLE_TRAINER):
            name = ALLE_TRAINER[i]
            ab   = name in abwesend.get("Trainer", [])
            set_cell(ws, row, trainer_col,
                     f"{'✗' if ab else '✓'} {name}",
                     FARBEN["abwesend_turner"] if ab else FARBEN["anwesend"],
                     font(size=9, color="FFFFFF"), align(h="left"))
        else:
            ws.cell(row=row, column=trainer_col).fill = fill(a)
        row += 1

    row += 1

    merge_set(ws, row, 2, end_col, "GERAETE-LEGENDE",
        FARBEN["legende_bg"], font(bold=True, size=10))
    row += 1
    items = [
        (geraet_1, FARBEN["g1_blau"]),
        (geraet_2, FARBEN["g2_orange"]),
        ("Aufwaermen", FARBEN["aufwaermen"]),
        ("Aufbauen/Abbauen", FARBEN["aufbauen"]),
        ("Springer", FARBEN["springer"]),
    ]
    for ci, (label, hex_c) in enumerate(items, start=2):
        set_cell(ws, row, ci, label, hex_c, font(bold=True, size=9), align(wrap=True))
    row += 1

    row += 1

    merge_set(ws, row, 2, end_col, "TRAINER-EINTEILUNG",
        FARBEN["legende_bg"], font(bold=True, size=10))
    row += 1

    # Nur anwesende Trainer in der Einteilung zeigen
    anwesende_trainer = [t for t in ALLE_TRAINER if t not in abwesend.get("Trainer", [])]

    set_cell(ws, row, 2, "Zeit", FARBEN["header"], font(bold=True))
    for ci, trainer in enumerate(anwesende_trainer, start=3):
        set_cell(ws, row, ci, trainer, FARBEN["header"], font(bold=True))
    row += 1

    for slot_idx, (row_start, row_end) in enumerate(grid_rows):
        slot_label = f"{_min_to_hhmm(row_start)}–{_min_to_hhmm(row_end)}"
        ws.row_dimensions[row].height = 38
        set_cell(ws, row, 2, slot_label, FARBEN["header"], font(bold=True))

        for ci, trainer in enumerate(anwesende_trainer, start=3):
            plan = trainer_plan.get(trainer)
            if plan is None or slot_idx >= len(plan):
                continue

            text, fk = plan[slot_idx]
            if slot_idx == 0 and trainer in sondertiming:
                text = sondertiming[trainer]
                fk   = "sonder"

            tc = "555555" if fk == "abwesend" else "FFFFFF"
            set_cell(ws, row, ci, text, FARBEN.get(fk, "FFFFFF"),
                     font(bold=True, color=tc), align(wrap=True))
        row += 1

    if anmerkungen:
        merge_set(ws, row, 2, end_col, "ANMERKUNGEN",
            FARBEN["anmerkung"], font(bold=True, size=9, color="1A5276"))
        row += 1
        for anm in anmerkungen:
            merge_set(ws, row, 2, end_col, anm,
                FARBEN["anmerkung"], font(size=9, color="1A5276"), align(h="left"))
            row += 1

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    out_dir   = "/tmp/trainingsplan"
    os.makedirs(out_dir, exist_ok=True)
    xlsx_path = os.path.join(out_dir, f"{datum_kurz}_Trainingsplan.xlsx")
    pdf_path  = os.path.join(out_dir, f"{datum_kurz}_Trainingsplan.pdf")
    wb.save(xlsx_path)
    print(f"[OK] Excel: {xlsx_path}")

    tmp_dir = tempfile.mkdtemp()
    result  = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "pdf",
         "--outdir", tmp_dir, xlsx_path],
        capture_output=True, text=True, timeout=90
    )
    if result.returncode == 0:
        tmp_pdf = os.path.join(tmp_dir,
            os.path.basename(xlsx_path).replace(".xlsx", ".pdf"))
        if os.path.exists(tmp_pdf):
            shutil.copy2(tmp_pdf, pdf_path)
            print(f"[OK] PDF: {pdf_path}")
        else:
            raise RuntimeError("LibreOffice PDF nicht gefunden")
    else:
        raise RuntimeError(f"LibreOffice Fehler: {result.stderr[:300]}")

    return xlsx_path, pdf_path

# ════════════════════════════════════════════════════════════════
#  WHATSAPP (CallMeBot)
# ════════════════════════════════════════════════════════════════

def send_email(text):
    import os, smtplib, ssl as _ssl
    from email.mime.text import MIMEText
    user = os.environ.get("GMAIL_USER", "")
    pw   = os.environ.get("GMAIL_APP_PASSWORD", "")
    to   = os.environ.get("EMAIL_TO", "") or user
    if not user or not pw:
        print("[MAIL] keine Gmail-Zugangsdaten - uebersprungen.")
        return
    try:
        lines = [l for l in text.strip().splitlines() if l.strip()]
        subj = "TV Rheinzabern: " + (lines[0][:80] if lines else "Info")
        msg = MIMEText(text, _charset="utf-8")
        msg["Subject"] = subj
        msg["From"] = user
        msg["To"] = to
        ctx = _ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx) as s:
            s.login(user, pw)
            s.sendmail(user, [to], msg.as_string())
        print("[MAIL] gesendet an", to)
    except Exception as e:
        print("[MAIL] Fehler:", e)


def send_whatsapp(text):
    """Sendet eine WhatsApp-Nachricht via CallMeBot (kostenlos)."""
    send_email(text)
    if not WHATSAPP_PHONE or not CALLMEBOT_APIKEY:
        print(f"[WA-TEST] {text[:200]}")
        return
    try:
        encoded = urllib.parse.quote(text)
        url = f"https://api.callmebot.com/whatsapp.php?phone={WHATSAPP_PHONE}&text={encoded}&apikey={CALLMEBOT_APIKEY}"
        with urllib.request.urlopen(url, timeout=15) as r:
            print(f"[OK] WhatsApp gesendet (HTTP {r.status}): {text[:80]}...")
    except Exception as e:
        print(f"[FEHLER] WhatsApp konnte nicht gesendet werden: {e}")

# ════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════

def apply_config_roster(sftp):
    """Laedt config/config.json (von admin.php gepflegt) und befuellt
    ALLE_TURNER, ALLE_TRAINER, WEBSITE_TO_DISPLAY, GRUPPEN_ORDER,
    GRUPPEN_ZEITEN, IMMER_SPRINGER. Seit Phase 2 (24.08.2026) gibt es KEINE
    hartkodierte Namensliste mehr im Repo -- bei jedem Fehler bleiben die
    Strukturen leer, und main() bricht danach ueber den
    REQUIRE_SERVER_ROSTER-Check klar ab, statt einen kaputten/leeren Plan zu
    erzeugen. `gruppen` kann sowohl das neue Objekt-Format
    ({"name":..,"zeiten":{...}}, seit dem Gruppenzeiten-Umbau) als auch das
    alte String-Format (vor der Migration) enthalten -- fehlt "zeiten",
    greifen die migrierten Default-Werte (_default_zeiten_for)."""
    global ALLE_TURNER, ALLE_TRAINER, WEBSITE_TO_DISPLAY, GRUPPEN_ORDER, GRUPPEN_ZEITEN, IMMER_SPRINGER, GRUPPEN_TAUSCH
    try:
        f = sftp.open("config/config.json", "r")
        cfg = json.loads(f.read().decode("utf-8", errors="replace"))
        f.close()
    except Exception as e:
        print(f"[CONFIG] config.json nicht ladbar ({e!r}).")
        return
    try:
        gruppen_raw = cfg.get("gruppen") or []
        gruppen_names, zeiten, tausch = [], {}, set()
        for g in gruppen_raw:
            if isinstance(g, dict) and g.get("name"):
                name = g["name"]
                gz = g.get("zeiten")
                zeiten[name] = gz if isinstance(gz, dict) and gz.get("mi") and gz.get("fr") else _default_zeiten_for(name)
                if g.get("geraet_tausch"):
                    tausch.add(name)
                gruppen_names.append(name)
            elif isinstance(g, str):
                zeiten[g] = _default_zeiten_for(g)
                gruppen_names.append(g)
        if not gruppen_names:
            gruppen_names = ["G1", "G2", "G3", "G4"]
            zeiten = {g: _default_zeiten_for(g) for g in gruppen_names}

        turner  = {g: [] for g in gruppen_names}
        trainer = []
        immer_springer = set()
        w2d     = {}
        for p in cfg.get("personen", []):
            ni = p.get("name_intern") or p.get("anzeige")
            if not ni:
                continue
            anz = (p.get("anzeige") or ni).strip()
            key = re.sub(r'\s*\(G(\d+)\)$', r' G\1', anz)
            w2d[key] = ni
            if p.get("rolle") == "trainer":
                trainer.append(ni)
                if p.get("immer_springer"):
                    immer_springer.add(ni)
            elif p.get("rolle") == "turner":
                g = p.get("gruppe")
                turner.setdefault(g, []).append(ni)
        if not trainer or not any(turner.values()):
            print("[CONFIG] config.json unvollstaendig.")
            return
        ALLE_TURNER        = turner
        ALLE_TRAINER       = trainer
        WEBSITE_TO_DISPLAY = w2d
        GRUPPEN_ORDER      = gruppen_names
        GRUPPEN_ZEITEN     = zeiten
        IMMER_SPRINGER     = immer_springer
        GRUPPEN_TAUSCH     = tausch
        print(f"[CONFIG] Roster aus config.json: "
              f"{sum(len(v) for v in turner.values())} Turner, {len(trainer)} Trainer, "
              f"{len(gruppen_names)} Gruppen, {len(immer_springer)} immer-Springer, "
              f"{len(tausch)} mit Geraet-Tausch.")
    except Exception as e:
        print(f"[CONFIG] Fehler beim Aufbau ({e!r}).")


def require_server_roster():
    """Harter Abbruch, wenn der Roster nicht vom Server geladen werden konnte.
    Seit Phase 2 gibt es keine hartkodierten Namen mehr im Code, die als
    Fallback einspringen koennten -- ein leerer/kein Server-Zustand darf NICHT
    stillschweigend zu einem falschen/leeren Trainingsplan fuehren, sondern
    muss den Lauf klar abbrechen (siehe Vault: GitHub-Umzug, Phase 2)."""
    if not GRUPPEN_ORDER or not any(ALLE_TURNER.values()) or not ALLE_TRAINER:
        msg = ("Trainingsplan-Lauf abgebrochen: Roster konnte nicht von config/config.json "
               "geladen werden (leer oder nicht erreichbar). Es gibt keinen Namens-Fallback "
               "im Code mehr (Phase 2, 24.08.2026), damit die Automatik nie mit falschen/"
               "leeren Daten weiterlaeuft.")
        print(f"[FATAL] {msg}")
        try:
            send_whatsapp(f"Hi, Cloude hier 🚨\n\n{msg}")
        except Exception:
            pass
        raise SystemExit(msg)


def build_admin_trainer_plan(absences, partial, grid_rows, grid_phase, trainer_roles_history=None, ki=None):
    """Admin-Plan: manuell im Admin-Bereich gesetzte Trainer-Zellen (fixed_trainer_partial)
    sind HARTE Vorgaben. Manuell belegte Trainer werden aus der Auto-Verteilung
    herausgenommen. Der REST wird um diese Vorgabe herum geplant - und zwar inklusive
    KI-/Kommando-Anweisungen (ki.assign/merges aus Anmerkungen), falls vorhanden, statt
    sie stillschweigend zu verwerfen.
    Bugfix 19.08.2026 (Noah: "im Admin-Bereich festgelegte Bereiche werden ignoriert,
    die KI soll aussenrum planen"): vorher wurde bei vorhandenem fixed_trainer_partial
    IMMER nur der reine Auto-Builder (build_trainer_plan, ohne KI) fuer den Rest benutzt
    - jede KI-Zuteilung/Zusammenlegung aus einer gleichzeitigen Anmerkung ging verloren.
    Jetzt: von Admin fest belegte Gruppen (aus dem Zellentext der committed Trainer
    erkannt) werden fuer den Rest-Builder als 'voll abwesend' markiert, damit sie nicht
    doppelt vergeben werden - und der Rest nutzt build_ki_einteilung, wenn ki.assign/
    merges vorhanden sind. Leere Randzeiten (erste/letzte Zeile) werden mit
    Aufbauen/Abbauen gefuellt."""
    partial = partial or {}
    ki = ki or {}
    committed = [t for t, s in partial.items()
                 if isinstance(s, list) and any((c and len(c) >= 2 and (c[0] or c[1])) for c in s)]
    nulled = [t for t, s in partial.items() if s is None]

    # Von den fest zugewiesenen Trainern belegte Gruppen ermitteln (aus dem
    # Zellentext, z.B. "G3+G4" oder "G1"), damit der Rest-Builder sie nicht
    # nochmal vergibt (sonst doppelte Gruppenbelegung moeglich).
    covered_groups = set()
    for t in committed:
        labels = {str(c[0]).strip() for c in partial[t]
                  if isinstance(c, (list, tuple)) and len(c) >= 2 and c[0]}
        for lab in labels:
            norm = _ki_norm_merge(lab) or (lab if lab in GRUPPEN_ORDER else None)
            if norm:
                covered_groups.update(norm.split("+"))

    abs2 = {k: list(v) for k, v in absences.items()}
    abs2.setdefault("Trainer", [])
    for t in committed + nulled:
        if t not in abs2["Trainer"]:
            abs2["Trainer"].append(t)
    for g in covered_groups & set(GRUPPEN_ORDER):
        existing = abs2.get(g, [])
        for name in ALLE_TURNER.get(g, []):
            if name not in existing:
                existing.append(name)
        abs2[g] = existing

    if ki.get("assign") or ki.get("merges"):
        try:
            base, _s, _a = build_ki_einteilung(abs2, ki, grid_rows, grid_phase, trainer_roles_history)
        except Exception as _e:
            print(f"[ADMIN] KI-Einteilung um Admin-Fix herum fehlgeschlagen ({_e!r}), Fallback Standard-Builder.")
            base, _s, _a = build_trainer_plan(abs2, grid_rows, grid_phase, trainer_roles_history=trainer_roles_history)
    else:
        base, _s, _a = build_trainer_plan(abs2, grid_rows, grid_phase, trainer_roles_history=trainer_roles_history)
    for t in committed:
        cells = partial[t]
        n = len(cells)
        out = []
        for i, c in enumerate(cells):
            if c and len(c) >= 2 and (c[0] or c[1]):
                out.append((c[0], c[1]))
            elif i == 0:
                out.append(("Aufbauen", "aufbauen"))
            elif i == n - 1:
                out.append(("Abbauen", "aufbauen"))
            else:
                out.append(("", ""))
        base[t] = out
    for t in nulled:
        base[t] = None
    return base



# ════════════════════════════════════════════════════════════════
#  MINI-KI: freie Trainer-Anmerkungen -> Planaenderungen (GitHub Models)
# ════════════════════════════════════════════════════════════════
GH_MODELS_ENDPOINT = os.environ.get("GH_MODELS_ENDPOINT", "https://models.github.ai/inference/chat/completions")
GH_MODELS_MODEL    = os.environ.get("GH_MODELS_MODEL", "openai/gpt-4o")

def _ki_token():
    return os.environ.get("GH_MODELS_TOKEN") or os.environ.get("GITHUB_TOKEN") or ""

def _ki_call(messages, token, timeout=60):
    body = json.dumps({"model": GH_MODELS_MODEL, "messages": messages,
                       "temperature": 0, "response_format": {"type": "json_object"}}).encode("utf-8")
    req = urllib.request.Request(GH_MODELS_ENDPOINT, data=body, method="POST")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Accept", "application/json")
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode("utf-8"))["choices"][0]["message"]["content"]

_KI_PROMPT = """Du wandelst eine freie Trainer-Anmerkung fuer einen Kinder-Turntrainingsplan in striktes JSON um.

Kontext:
- Gruppen: <GR>
- Trainer: <TR>
- Geraete: <GE>
- Heute/naechstes Training: <NEXT>
- Diese Anmerkung stammt vom Trainer: "<WRITER>". "ich"/"mich"/"mir"/"meine" bezieht sich immer auf diesen Trainer.

Aktuelle Anmerkungen je Datum — NUR Referenz fuer Text-Aenderungen, leite daraus NIEMALS Aktionen (gruppe_entfall/timing/zuteilung/geraete/merges) ab: <CURRENT>

Gib GENAU dieses JSON zurueck (keine weiteren Felder, kein Text drumherum):
{"ziel": {"typ": "naechstes | datum | naechstes_geraet", "datum": "TT.MM.JJJJ oder null", "geraet": "Geraetename oder null"}, "notiz": "reiner Hinweistext oder null", "geraete": {"g1": "Geraet", "g2": "Geraet"} , "gruppe_entfall": ["G3"], "abwesend_kinder": [], "wieder_da_kinder": [], "abwesend_trainer": [], "einteilung": {}, "timing": [{"trainer": "Name", "richtung": "spaet | frueh", "uhrzeit": "HH:MM oder null"}], "merges": [["G3","G4"]], "zuteilung": [{"trainer": "Name", "gruppe": "z.B. G1, G2+G3 oder Springer"}], "notiz_neu": null, "reset": false, "unsicher": true}

Regeln:
- ziel.typ: KEIN Datum/Zeitpunkt genannt -> "naechstes" (gilt NUR fuer genau das naechste Training, NIE Dauerauftrag). Konkretes Datum (z.B. "am 14.10") -> "datum" mit datum. "naechstes Mal <Geraet>" / "wenn wir <Geraet> turnen" -> "naechstes_geraet" mit geraet. Wenn ziel unklar -> typ "naechstes".
- "kommt spaeter"/"erst um X"/"ist ab X (wieder) da" -> timing richtung "spaet"; "geht frueher"/"muss um X weg/los"/"geht ab X"/"nur bis X da"/"ist ab X weg" -> "frueh". uhrzeit als HH:MM oder null. WICHTIG: jede Nennung einer Uhrzeit zusammen mit einem Trainer, der teilweise da ist, IMMER als timing eintragen, NIEMALS als abwesend_trainer -- abwesend_trainer ist ausschliesslich fuer einen Trainer, der am GESAMTEN Training gar nicht da ist.
- "GX entfaellt"/"faellt aus" -> gruppe_entfall ["GX"].
- Geraetewahl "wir turnen X und Y" -> geraete {"g1":X,"g2":Y}, sonst geraete null.
- Gruppen zusammenlegen OHNE genannten Trainer (z.B. "G3 und G4 zusammenlegen") -> merges [["G3","G4"]]. MIT genanntem Trainer ("Fabian macht G1+G2") -> zuteilung {trainer, gruppe:"G1+G2"}. "ich mache ..." -> trainer = der Schreiber. gruppe darf "Springer" oder "GX+GY" sein.
- Reiner Hinweis ("schreibe/bitte ...") -> nur notiz.
- Nur Namen aus der Trainerliste, nur Gruppen aus der Gruppenliste, nur Geraete aus der Geraeteliste.
- Keine Aktion in einem Feld: leere Liste [] bzw. null. Bei Unklarheit unsicher=true und notiz=Originaltext.
- Ist ein Trainer NICHT namentlich genannt (z.B. "der andere Trainer", "jemand", "ein Trainer macht G2+G3") -> als merges eintragen (ohne Trainer), NICHT als zuteilung.
- "loesche/entferne alle (bisherigen) Anmerkungen", "setze zurueck", "reset", "mach alles rueckgaengig" -> reset: true (alle bisherigen Vorgaben fuer das Ziel-Datum werden entfernt).
- notiz_neu: Setze es bei JEDER Aenderung des Anmerkungs-TEXTES — hinzufuegen ("fuege hinzu ..."), loeschen ("loesche die Zeile ..."), umformulieren oder ersetzen ("aendere X zu Y"). Gib dann IMMER den KOMPLETTEN neuen Anmerkungstext des Ziel-Datums zurueck: bestehende Zeilen aus der Referenz uebernehmen und die Aenderung anwenden. Bei reinen Struktur-/Plan-Aktionen ohne Text-Aenderung notiz_neu = null.
- WICHTIG: gruppe_entfall/timing/zuteilung/merges/geraete/reset NUR aus der AKTUELLEN Trainer-Anmerkung ableiten, NIEMALS aus der Referenz der aktuellen Anmerkungen.
- abwesend_kinder / wieder_da_kinder: einzelne Kinder ab-/wieder anmelden (nur Namen aus der Kinderliste). abwesend_trainer: NUR Trainer, die das GESAMTE Training ueber fehlen (Urlaub, krank, Spaetschicht o.ae.) -- niemals fuer einen Trainer, der nur teilweise da ist, auch wenn "fehlt ab X" o.ae. klingt wie eine Abwesenheit (siehe timing-Regel oben).
- einteilung: KOMPLETTE Zuordnung {Trainer: Label}, wenn die Anweisung Trainer den Gruppen zuordnet (z.B. "Trainer A Gruppe 1, Trainer B Gruppe 2, Rest Springer"). Label = eine der konfigurierten Gruppen, "GX+GY" (mehrere Gruppen durch + verbunden) oder "Springer". Sonst {} = automatische Einteilung.
- Antworte ausschliesslich mit dem JSON-Objekt."""

def _ki_system_prompt(writer, next_date_str, current_notes=None):
    gr = "; ".join(f"{g}: {', '.join(ALLE_TURNER[g])}" for g in ALLE_TURNER)
    tr = ", ".join(ALLE_TRAINER)
    ge = ", ".join(sorted({g for c in GERAETE_ROTATION for g in c}))
    cur = "(keine)"
    if current_notes:
        cur = " | ".join(f"{d}: {(t or '').replace(chr(10), ' / ')}" for d, t in current_notes.items()) or "(keine)"
    return (_KI_PROMPT.replace("<GR>", gr).replace("<TR>", tr).replace("<GE>", ge)
            .replace("<NEXT>", next_date_str).replace("<WRITER>", writer or "?").replace("<CURRENT>", cur))

def ki_analyze(notiz, writer, next_date_str, token, current_notes=None):
    content = _ki_call([{"role": "system", "content": _ki_system_prompt(writer, next_date_str, current_notes)},
                        {"role": "user", "content": notiz}], token)
    return json.loads(content)

def _ki_full_name(short):
    s = (short or "").strip()
    if s in ALLE_TRAINER:
        return s
    return WEBSITE_TO_DISPLAY.get(s, s)

def _ki_time_min(s):
    m = re.search(r'(\d{1,2})[:.](\d{2})', s or "")
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if 0 <= h < 24 and 0 <= mi < 60:
            return h * 60 + mi
    return None

def ki_timing_to_dict(timing_list, grid_rows):
    out = {}
    for tm in (timing_list or []):
        name = _ki_full_name(tm.get("trainer", ""))
        if name not in ALLE_TRAINER:
            continue
        kind = "spaet" if tm.get("richtung") == "spaet" else "frueh"
        t = _ki_time_min(tm.get("uhrzeit"))
        blocked = compute_blocked_slots(kind, t, grid_rows)
        out[name] = {"kind": kind, "time_min": t,
                     "time_str": (f"{t//60:02d}:{t%60:02d}" if t is not None else None),
                     "notiz": "", "blocked": sorted(blocked)}
    return out

def _ki_next_training_after(d):
    d = d + timedelta(days=1)
    while d.weekday() not in (2, 4):
        d += timedelta(days=1)
    return d

def ki_resolve_target(ziel, state):
    ziel = ziel or {}
    typ = ziel.get("typ") or "naechstes"
    if typ == "datum" and ziel.get("datum"):
        for fmt in ("%d.%m.%Y", "%d.%m.%y"):
            try:
                return datetime.strptime(ziel["datum"], fmt).date().strftime("%d.%m.%y")
            except Exception:
                pass
    if typ == "naechstes_geraet" and ziel.get("geraet"):
        g = ziel["geraet"]
        try:
            idx, _ = get_next_gear(state, exclude_date=None)
        except Exception:
            idx = state.get("geraet_combo_index", 0)
        d = active_training_date()
        for _ in range(24):
            if g in GERAETE_ROTATION[idx % len(GERAETE_ROTATION)]:
                return d.strftime("%d.%m.%y")
            d = _ki_next_training_after(d)
            idx += 1
    return active_training_date().strftime("%d.%m.%y")

# ---- KI-Einteilung: Zusammenlegen / Zuteilen im Raster ----

def _ki_norm_merge(label):
    parts = [p for p in re.split(r'\+', str(label or "").replace(" ", "")) if p in GRUPPEN_ORDER]
    if len(parts) < 2:
        return None
    return "+".join(sorted(parts, key=lambda g: GRUPPEN_ORDER.index(g)))

def _merge_small_singletons(groups, present, gruppen_zeiten, min_kids=3):
    """Legt zu kleine (<min_kids anwesende Kinder) benachbarte Einzelgruppen
    zusammen -- nur wenn sie zeit-kompatibel sind (siehe zeiten_kompatibel).
    Kein kompatibler Nachbar da -> Gruppe bleibt einzeln (Notlage-Fallback
    statt eines unmoeglichen Merges)."""
    units = [[g] for g in groups]
    def cnt(u): return sum(present.get(g, 0) for g in u)
    def compatible(a, b): return all(zeiten_kompatibel(gruppen_zeiten, x, y) for x in a for y in b)
    changed = True
    while changed:
        changed = False
        for i, u in enumerate(units):
            if cnt(u) < min_kids and len(units) > 1:
                cands = [j for j in (i - 1, i + 1) if 0 <= j < len(units) and compatible(u, units[j])]
                if not cands:
                    continue
                best = min(cands, key=lambda j: cnt(units[j]) + cnt(u))
                lo, hi = sorted((i, best))
                units[lo] = units[lo] + units[hi]; del units[hi]
                changed = True; break
    return units


def build_ki_einteilung(absences, ki, grid_rows, grid_phase, trainer_roles_history=None):
    """Isolierter Builder aus KI-Anweisungen (cancel/merges/assign) -> vollstaendiges
    Raster. Entfallene Gruppen -> kein eigener Trainer (Trainer wird Springer).
    Nutzt dieselbe Fairness-Funktion (_assign_units_fair) und denselben
    Zellen-Builder (_cells_for_unit) wie der Standard-Pfad build_trainer_plan,
    damit beide Pfade nie auseinanderlaufen koennen (siehe Farb-Kollisions-
    Historie 12./19.08.2026 -- genau diese Art Duplizierung war die Ursache)."""
    trainer_roles_history = trainer_roles_history or {}
    abwesend = set(absences.get("Trainer", []))
    available = [t for t in ALLE_TRAINER if t not in abwesend]
    cancel = set(g for g in (ki.get("cancel") or []) if g in GRUPPEN_ORDER)
    for g in GRUPPEN_ORDER:
        tl = ALLE_TURNER.get(g, [])
        if tl and all(t in absences.get(g, []) for t in tl):
            cancel.add(g)
    base_groups = [g for g in GRUPPEN_ORDER if g not in cancel]

    def _compat(grps):
        return all(zeiten_kompatibel(GRUPPEN_ZEITEN, x, y) for x in grps for y in grps)

    assign_raw = ki.get("assign") or []
    used_groups = set()
    merge_units = []   # (groups_set, forced_trainer_or_None)
    for a in assign_raw:
        tr = _ki_full_name(a.get("trainer", ""))
        lab = _ki_norm_merge(a.get("gruppe") or a.get("label") or "")
        if tr in available and lab:
            grps = set(lab.split("+")) & set(base_groups)
            if len(grps) >= 2 and not (grps & used_groups) and _compat(grps):
                merge_units.append((grps, tr)); used_groups |= grps
    for mg in (ki.get("merges") or []):
        grps = set(g for g in (mg or []) if g in base_groups) - used_groups
        if len(grps) >= 2 and _compat(grps):
            merge_units.append((grps, None)); used_groups |= grps

    forced = {}
    for a in assign_raw:
        tr = _ki_full_name(a.get("trainer", ""))
        rawlab = (a.get("gruppe") or a.get("label") or "").strip()
        if tr not in available or "+" in rawlab:
            continue
        if rawlab.lower() == "springer":
            forced[tr] = "Springer"
        elif rawlab in base_groups and rawlab not in used_groups:
            forced[tr] = rawlab; used_groups.add(rawlab)

    units = []   # (label, forced_trainer_or_None, groups_list)
    for grps, tr in merge_units:
        glist = sorted(grps, key=lambda g: GRUPPEN_ORDER.index(g))
        units.append((unit_label(glist), tr, glist))
    # Uebrig gebliebene Einzelgruppen: zu kleine Gruppen (<3 anwesende Kinder)
    # automatisch mit einem zeit-kompatiblen Nachbarn zusammenlegen (wie im
    # Standard-Pfad build_trainer_plan).
    present = {g: max(0, len(ALLE_TURNER.get(g, [])) - len(absences.get(g, [])))
               for g in GRUPPEN_ORDER}
    leftover = [g for g in base_groups if g not in used_groups]
    for grp_list in _merge_small_singletons(leftover, present, GRUPPEN_ZEITEN):
        units.append((unit_label(grp_list), None, grp_list))
        used_groups.update(grp_list)

    unit_groups = {lab: grps for lab, _tr, grps in units}
    assign = dict(forced)
    for lab, tr, _grps in units:
        if tr:
            assign[tr] = lab
    pool = [t for t in available if t not in assign]
    open_units = [lab for lab, tr, _g in units if not tr]

    fair_assign, springers = _assign_units_fair(open_units, pool, trainer_roles_history, IMMER_SPRINGER)
    assign.update(fair_assign)

    TRAINER_PLAN = {}
    for t in ALLE_TRAINER:
        if t in assign:
            lab = assign[t]
            if lab == "Springer":
                TRAINER_PLAN[t] = _springer_cells(grid_rows)
            else:
                groups = unit_groups.get(lab) or [p for p in lab.split("+") if p in GRUPPEN_ORDER] or [lab]
                TRAINER_PLAN[t] = _cells_for_unit(lab, groups, grid_rows, grid_phase)
        elif t in springers:
            TRAINER_PLAN[t] = _springer_cells(grid_rows)
        else:
            TRAINER_PLAN[t] = None
    anm = []
    unfilled = [lab for lab in open_units if lab not in fair_assign]
    if unfilled:
        anm.append("ACHTUNG: kein Trainer mehr fuer: " + ", ".join(unfilled))
    return TRAINER_PLAN, {}, anm

def _ki_kid(name):
    disp = normalize_name((name or "").strip())
    for g, lst in ALLE_TURNER.items():
        if disp in lst:
            return g, disp
    return None

def _merge_ki_assign(existing, new):
    """Fuegt neue Zuteilungen zu bestehenden hinzu, statt sie zu ueberschreiben.
    Spaetere Eintraege fuer denselben Trainer gewinnen; andere Trainer bleiben erhalten.
    (Noah 16.07.2026: sonst warf eine zweite Anmerkung die erste raus, z.B. 'Noah Springer'.)"""
    by = {}
    order = []
    for a in list(existing or []) + list(new or []):
        if not isinstance(a, dict):
            continue
        tr = (a.get("trainer") or "").strip()
        gr = a.get("gruppe") or a.get("label")
        if not tr or not gr:
            continue
        if tr not in by:
            order.append(tr)
        by[tr] = gr
    return [{"trainer": t, "gruppe": by[t]} for t in order]


# ════════════════════════════════════════════════════════════════
#  DETERMINISTISCHER ANMERKUNGS-KOMMANDO-PARSER (Bug-Fix 19.08.2026)
# ════════════════════════════════════════════════════════════════
# Trainer schreiben Kurz-Kommandos wie "Noah Springer" oder "lösche alle
# Anmerkungen" in die Anmerkungs-Zeile - vorher landete das 1:1 als Text im
# naechsten Plan (statt als Anweisung ausgefuehrt zu werden), weil die
# KI-Interpretation entweder keinen GH_MODELS_TOKEN hatte oder scheiterte.
# Dieser Parser erkennt eine feste Liste kanonischer Muster deterministisch
# UND OHNE KI, wendet die Aktion direkt auf fixed_entries[Ziel-Datum] an und
# konsumiert die Anmerkung (markiert sie als gelesen). Nur wenn ALLE Zeilen
# einer Anmerkung sicher matchen, wird sie hier verarbeitet - sonst bleibt
# sie der KI/dem Copy-Paste-Fallback ueberlassen.
#
# Erkannte Muster (case-insensitive, jeweils EINE Zeile):
#   - "lösche/loesche/entferne (alle) (bisherigen) Anmerkungen"
#     "alle Anmerkungen löschen" / "reset" / "zurücksetzen"
#     -> reset: alle Vorgaben fuer das Ziel-Datum werden entfernt
#   - "<Trainer-Vorname> Springer"       (z.B. "Noah Springer")
#   - "<Trainer-Vorname> G1" / "G2" / "G3" / "G4"
#   - "<Trainer-Vorname> G1+G2"          (auch "G3+G4", "Gruppe 1+2", ...)
#   - "<Trainer-Vorname> abwesend/nicht da/fehlt"
#   - "<Trainer-Vorname> anwesend/wieder da/kommt"
#   - "GX entfaellt" / "Gruppe X entfaellt"
#   - "GX und GY zusammenlegen" / "GX+GY zusammenlegen"
#
# Erweiterbar: neues Regex + Handler unten in COMMAND_PATTERNS ergaenzen.

# Trainer-Vornamen und Gruppennamen kommen seit dem Gruppenzeiten-Umbau
# ausschliesslich aus der zur Laufzeit geladenen Config (ALLE_TRAINER/
# GRUPPEN_ORDER) -- keine hartkodierte Namensliste mehr im (oeffentlichen)
# Repo. Die Regexe werden deshalb erst NACH apply_config_roster() bei Bedarf
# gebaut (Aufrufhaeufigkeit: wenige Anmerkungen pro Lauf, unkritisch).

def _first_to_full(first):
    for full in ALLE_TRAINER:
        if full.split()[0].lower() == first.lower():
            return full
    return None

def _resolve_group_name(tok):
    """Loest ein rohes Token gegen die aktuell konfigurierten Gruppennamen
    auf (case-insensitive). Gruppen sind frei benennbar (Noah kann sie im
    Admin-Bereich umbenennen/entfernen), deshalb keine 'GN'-Digit-Annahme."""
    lower_map = {g.lower(): g for g in GRUPPEN_ORDER}
    return lower_map.get((tok or "").strip().lower())

def _name_alt(names):
    ordered = sorted({n for n in names if n}, key=len, reverse=True)
    return "|".join(re.escape(n) for n in ordered)

_RE_RESET = re.compile(
    r'^\s*(?:'
    r'(?:l(?:oe|ö)sch(?:e|t|en)?|entfern(?:e|t|en)?)\s+(?:alle\s+)?(?:bisherigen\s+)?anmerkungen'
    r'|alle\s+anmerkungen\s+(?:l(?:oe|ö)schen|entfernen)'
    r'|reset|zur(?:ue|ü)cksetzen|mach\s+alles\s+r(?:ue|ü)ckg(?:ae|ä)ngig'
    r')\s*[.!]?\s*$',
    re.IGNORECASE,
)

def _re_trainer(suffix_pattern):
    """Baut 'ein Trainer-Vorname gefolgt von <suffix_pattern>' aus der aktuell
    geladenen Trainerliste. None, wenn (noch) kein Roster geladen ist."""
    firsts = _name_alt(t.split()[0] for t in ALLE_TRAINER if t)
    if not firsts:
        return None
    return re.compile(r'^\s*(?P<name>' + firsts + r')' + suffix_pattern + r'\s*[.!]?\s*$', re.IGNORECASE)

def _re_trainer_role():
    return _re_trainer(r'\s+(?:macht\s+|ist\s+|:\s*)?(?P<lab>Springer|[\wÄÖÜäöüß]+(?:\s*\+\s*[\wÄÖÜäöüß]+)*)')

def _re_trainer_absent():
    return _re_trainer(r'\s+(?:ist\s+)?(?:abwesend|nicht\s+da|fehlt|kann\s+nicht|f(?:ae|ä)llt\s+aus)')

def _re_trainer_present():
    return _re_trainer(r'\s+(?:ist\s+)?(?:anwesend|wieder\s+da|kommt(?:\s+doch)?|doch\s+da)')

def _re_trainer_timing():
    """Trainer-Name gefolgt von beliebigem Resttext -- der Resttext wird
    danach mit derselben parse_trainer_timing()/is_timing_note()-Erkennung
    geprueft, die auch fuer die strukturierten Abmeldungen gilt (siehe
    get_absences()). Bewusst als letzter Versuch NACH allen spezifischeren
    Mustern (assign/abwesend/anwesend/entfall/merge): ein Text, der schon
    eines der spezifischeren Muster trifft, kommt hier nie an, ein Text ohne
    erkennbare Verspaetungs-/Frueher-gehen-Woerter bleibt weiterhin fuer die
    KI liegen (is_timing_note() liefert dann False, siehe _parse_command_line)."""
    return _re_trainer(r'\s+(?P<rest>.+)')

_RE_GRUPPE_ENTFALL = re.compile(
    r'^\s*(?P<g>[\wÄÖÜäöüß]+)\s+(?:entf(?:ae|ä)llt|f(?:ae|ä)llt\s+aus)\s*[.!]?\s*$',
    re.IGNORECASE,
)

_RE_MERGE = re.compile(
    r'^\s*(?P<a>[\wÄÖÜäöüß]+?)\s*(?:\+|und|u\.)\s*(?P<b>[\wÄÖÜäöüß]+?)'
    r'\s+(?:zusammenlegen|zusammen|gemeinsam)\s*[.!]?\s*$',
    re.IGNORECASE,
)

def _norm_group_label(lab):
    """'G1', 'g1+g2', 'Springer' -> kanonisch, gegen GRUPPEN_ORDER aufgeloest
    (keine 'GN'-Digit-Annahme mehr -- Gruppen sind frei benennbar)."""
    s = lab.strip()
    if re.match(r'^\s*springer\s*$', s, re.IGNORECASE):
        return "Springer"
    parts = re.split(r'\s*\+\s*', s)
    out = []
    for p in parts:
        real = _resolve_group_name(p)
        if not real:
            return None
        out.append(real)
    if not out:
        return None
    if len(out) == 1:
        return out[0]
    return "+".join(sorted(out, key=lambda g: GRUPPEN_ORDER.index(g)))

def _parse_command_line(line):
    """Versucht, EINE Anmerkungs-Zeile als bekanntes Kommando zu erkennen.
    Gibt ein dict {typ, ...} zurueck oder None."""
    s = (line or "").strip().rstrip(",;")
    if not s:
        return None
    if _RE_RESET.match(s):
        return {"typ": "reset"}
    re_role = _re_trainer_role()
    m = re_role.match(s) if re_role else None
    if m:
        full = _first_to_full(m.group("name"))
        lab = _norm_group_label(m.group("lab"))
        if full and lab:
            return {"typ": "assign", "trainer": full, "gruppe": lab}
    re_abs = _re_trainer_absent()
    m = re_abs.match(s) if re_abs else None
    if m:
        full = _first_to_full(m.group("name"))
        if full:
            return {"typ": "trainer_abwesend", "trainer": full}
    re_pres = _re_trainer_present()
    m = re_pres.match(s) if re_pres else None
    if m:
        full = _first_to_full(m.group("name"))
        if full:
            return {"typ": "trainer_anwesend", "trainer": full}
    m = _RE_GRUPPE_ENTFALL.match(s)
    if m:
        g = _resolve_group_name(m.group("g"))
        if g:
            return {"typ": "gruppe_entfall", "gruppe": g}
    m = _RE_MERGE.match(s)
    if m:
        a, b = _resolve_group_name(m.group("a")), _resolve_group_name(m.group("b"))
        if a and b and a != b:
            return {"typ": "merge", "gruppen": sorted([a, b], key=lambda g: GRUPPEN_ORDER.index(g))}
    # Verspaetung/Frueher-gehen MIT eindeutiger Ziffern-Uhrzeit (z.B. "Cassi
    # geht ab 18:30", "Cassi muss um 18 Uhr los"): bisher wurde so etwas NIE
    # deterministisch erkannt, sondern immer der KI ueberlassen -- schlug die
    # KI fehl (kein Token) oder klassifizierte falsch als Abwesenheit statt
    # Timing, wurde der Trainer faelschlich komplett aus dem Plan entfernt
    # (Bug 27.08.2026, siehe Kommentar in build_trainer_plan).
    #
    # WICHTIG (Noah 28.08.2026, zu Recht): das hier ist NUR ein kostenloser
    # Schnellpfad fuer den eindeutigen Digit-Fall, kein Ersatz fuer die KI.
    # Wird KEINE Ziffern-Uhrzeit gefunden -- z.B. "so gegen halb sieben",
    # "nach dem Abendessen" -- wird bewusst NICHTS deterministisch angewendet
    # (return None), sondern die Notiz bleibt fuer die KI liegen. Eine
    # fruehere Fassung wandte hier bei erkanntem Schluesselwort aber fehlender
    # Uhrzeit trotzdem "kein Uhrzeit" an (blockierte dann nur die letzte
    # Rasterzeile) -- das haette echte, nur nicht als Ziffer geschriebene
    # Uhrzeiten (die die KI durchaus verstehen und normalisieren kann)
    # stillschweigend falsch/unvollstaendig behandelt statt sie der KI zu
    # ueberlassen. Nutzt fuer die Ziffern-Erkennung dieselbe Funktion wie die
    # strukturierten Abmeldungen (get_absences()), damit beide Wege konsistent
    # bleiben.
    re_time = _re_trainer_timing()
    m = re_time.match(s) if re_time else None
    if m:
        full = _first_to_full(m.group("name"))
        rest = m.group("rest")
        # Wird noch ein ANDERER Trainer erwaehnt (z.B. "Cassi geht ab 18:30,
        # Fabian macht G3+G4"), ist das vermutlich eine zusammengesetzte
        # Anweisung mit mehreren Anteilen -- das greedy rest-Muster wuerde den
        # zweiten Teil sonst stillschweigend in Cassis Notiz verschlucken statt
        # ihn anzuwenden. Lieber komplett der KI ueberlassen (die den ganzen
        # Satz auf einmal sieht), statt nur die Haelfte umzusetzen.
        others = _name_alt(t.split()[0] for t in ALLE_TRAINER if t and _first_to_full(t.split()[0]) != full)
        if full and others and re.search(r'\b(?:' + others + r')\b', rest, re.IGNORECASE):
            return None
        kind, tmin = parse_trainer_timing(rest)
        if full and tmin is not None and is_timing_note(rest, kind, tmin):
            if kind is None:
                kind = "spaet"
            tstr = f"{tmin//60:02d}:{tmin%60:02d}"
            return {"typ": "trainer_timing", "trainer": full, "kind": kind, "time_str": tstr}
    return None

def _apply_command(cmd, fe):
    """Wendet ein geparstes Kommando auf einen fixed_entries[dk]-Eintrag an.
    fe wird in-place mutiert. Setzt manuell_bearbeitet=True damit der Admin-
    Pfad im main() den Trainer-Plan aus diesen Vorgaben neu baut."""
    typ = cmd.get("typ")
    if typ == "reset":
        # Alle Vorgaben fuer dieses Datum entfernen (analog KI reset: true)
        fe.clear()
        return
    fe.setdefault("manuell_bearbeitet", True)
    fe["quelle"] = fe.get("quelle") or "cmd"
    ki = fe.setdefault("ki", {})
    if typ == "assign":
        existing = ki.get("assign") or []
        existing = [x for x in existing if x.get("trainer") != cmd["trainer"]]
        existing.append({"trainer": cmd["trainer"], "gruppe": cmd["gruppe"]})
        ki["assign"] = existing
    elif typ == "merge":
        ex = [tuple(sorted(m, key=lambda g: GRUPPEN_ORDER.index(g) if g in GRUPPEN_ORDER else 99))
              for m in (ki.get("merges") or [])]
        new = tuple(cmd["gruppen"])
        if new not in ex:
            ex.append(new)
        ki["merges"] = [list(m) for m in ex]
    elif typ == "gruppe_entfall":
        ki["cancel"] = sorted(set((ki.get("cancel") or []) + [cmd["gruppe"]]))
        fa = fe.setdefault("fixed_absences", {})
        g = cmd["gruppe"]
        fa.setdefault(g, [])
        for t in ALLE_TURNER.get(g, []):
            if t not in fa[g]:
                fa[g].append(t)
    elif typ == "trainer_abwesend":
        fa = fe.setdefault("fixed_absences", {})
        fa.setdefault("Trainer", [])
        if cmd["trainer"] not in fa["Trainer"]:
            fa["Trainer"].append(cmd["trainer"])
    elif typ == "trainer_anwesend":
        fa = fe.setdefault("fixed_absences", {})
        if cmd["trainer"] in fa.get("Trainer", []):
            fa["Trainer"].remove(cmd["trainer"])
    elif typ == "trainer_timing":
        # Landet in derselben Struktur wie eine KI-erkannte Verspaetung
        # (ki["timing"], siehe ki_timing_to_dict()) -- ein Trainer bleibt
        # dadurch eingeteilt (NICHT abwesend), nur seine Randzeiten werden
        # spaeter rot geblockt und an einen freien Trainer uebergeben.
        existing = [x for x in (ki.get("timing") or []) if x.get("trainer") != cmd["trainer"]]
        existing.append({"trainer": cmd["trainer"], "richtung": cmd["kind"], "uhrzeit": cmd["time_str"]})
        ki["timing"] = existing

def preprocess_deterministic_annotations(sftp, anmerkungen_server, fixed_entries):
    """Erkennt deterministisch parsbare Kommando-Anmerkungen, wendet sie auf
    fixed_entries[naechstes_Datum] an und konsumiert sie (markiert gelesen).
    Laeuft VOR ki_process_annotations - so klappt die Interpretation der
    einfachen Faelle auch ohne GH_MODELS_TOKEN. Alles was nicht sauber matcht,
    bleibt fuer die KI/den Copy-Paste-Pfad liegen."""
    if not anmerkungen_server:
        return
    dk = active_training_date().strftime("%d.%m.%y")
    changed = False
    processed_ids = []
    consumed = []
    for a in list(anmerkungen_server):
        notiz = (a.get("notiz") or "").strip()
        if not notiz:
            continue
        lines = [ln for ln in notiz.splitlines() if ln.strip()]
        if not lines:
            continue
        parsed = [_parse_command_line(ln) for ln in lines]
        if any(p is None for p in parsed):
            continue  # gemischt -> lieber der KI ueberlassen
        # Alle Zeilen sind Kommandos -> anwenden
        fe = fixed_entries.setdefault(dk, {})
        for p in parsed:
            _apply_command(p, fe)
        # Nach reset() kann fe leer sein -> Eintrag komplett verwerfen
        if not fe:
            fixed_entries.pop(dk, None)
        changed = True
        consumed.append(a)
        if a.get("id"):
            processed_ids.append(a["id"])
        writer = a.get("trainer", "?")
        print(f"[CMD] '{notiz[:60]}' ({writer}) -> {dk}  {[p['typ'] for p in parsed]}")
    for a in consumed:
        try:
            anmerkungen_server.remove(a)
        except ValueError:
            pass
    if changed:
        try:
            f = sftp.open("fixed_entries.json", "wb")
            f.write(json.dumps(fixed_entries, indent=2, ensure_ascii=False).encode("utf-8"))
            f.close()
            print("[CMD] fixed_entries.json aktualisiert.")
        except Exception as e:
            print(f"[CMD] fixed_entries Upload fehlgeschlagen: {e}")
    if processed_ids:
        try:
            mark_anmerkungen_gelesen(sftp, processed_ids)
        except Exception as e:
            print(f"[CMD] mark gelesen fehlgeschlagen: {e}")


def ki_process_annotations(sftp, anmerkungen_server, fixed_entries, state):
    """Analysiert jede ungelesene Anmerkung per KI und schreibt die Aenderung in
    fixed_entries[Zieldatum] (Admin-Mechanismus). Einmalig pro Datum."""
    token = _ki_token()
    if not token or not anmerkungen_server:
        return
    next_date_str = active_training_date().strftime("%d.%m.%Y")
    allowed_ger = {g for c in GERAETE_ROTATION for g in c}
    processed_ids = []
    changed = False
    current_notes = {d: (fe.get("notiz") or "") for d, fe in fixed_entries.items()
                     if isinstance(fe, dict) and (fe.get("notiz") or "").strip()}
    for a in list(anmerkungen_server):
        notiz = (a.get("notiz") or "").strip()
        writer = a.get("trainer", "")
        if not notiz:
            continue
        try:
            res = ki_analyze(notiz, writer, next_date_str, token, current_notes)
        except Exception as e:
            print(f"[KI] Analyse fehlgeschlagen ({writer}): {e}")
            continue
        try:
            dk = ki_resolve_target(res.get("ziel"), state)
            if res.get("reset"):
                fixed_entries.pop(dk, None)
                if a.get("id"):
                    processed_ids.append(a["id"])
                if a in anmerkungen_server:
                    anmerkungen_server.remove(a)
                changed = True
                print(f"[KI] RESET {dk} (alle Vorgaben entfernt)")
                continue
            fe = fixed_entries.setdefault(dk, {})
            fe["manuell_bearbeitet"] = True
            fe["quelle"] = "ki"
            ki = fe.setdefault("ki", {})
            _nn = res.get("notiz_neu")
            if _nn is not None:
                fe["notiz"] = str(_nn).strip()   # reine Anmerkungs-Textbearbeitung (Struktur bleibt)
                if a.get("id"):
                    processed_ids.append(a["id"])
                if a in anmerkungen_server:
                    anmerkungen_server.remove(a)
                changed = True
                print(f"[KI] Anmerkungs-Text bearbeitet -> {dk}")
                continue
            note_lines = []
            ge = res.get("geraete")
            if ge and ge.get("g1") in allowed_ger and ge.get("g2") in allowed_ger:
                fe["geraet_1"] = ge["g1"]; fe["geraet_2"] = ge["g2"]
            fa = fe.setdefault("fixed_absences", {})
            ent = [g for g in (res.get("gruppe_entfall") or []) if g in ALLE_TURNER]
            if ent:
                ki["cancel"] = sorted(set((ki.get("cancel") or []) + ent))
                for g in ent:
                    fa.setdefault(g, [])
                    for t in ALLE_TURNER[g]:
                        if t not in fa[g]:
                            fa[g].append(t)
                    note_lines.append(f"{g} entfaellt")
            if res.get("timing"):
                ki["timing"] = res["timing"]
                for tm in res["timing"]:
                    ri = "kommt spaeter" if tm.get("richtung") == "spaet" else "geht frueher"
                    uh = tm.get("uhrzeit")
                    note_lines.append(f"{tm.get('trainer','')} {ri}" + (f" {uh}" if uh else ""))
            if res.get("zuteilung"):
                ki["assign"] = _merge_ki_assign(ki.get("assign"), res["zuteilung"])
                for z in res["zuteilung"]:
                    if z.get("trainer") and z.get("gruppe"):
                        note_lines.append(f"{z['trainer']}: {z['gruppe']}")
            if res.get("merges"):
                ki["merges"] = res["merges"]
                for mg in res["merges"]:
                    if isinstance(mg, list) and len(mg) >= 2:
                        note_lines.append("+".join(mg) + " zusammen")
            for kn in (res.get("abwesend_kinder") or []):
                hit = _ki_kid(kn)
                if hit:
                    g, disp = hit
                    fa.setdefault(g, [])
                    if disp not in fa[g]:
                        fa[g].append(disp)
                    note_lines.append(f"{disp} abgemeldet")
            for kn in (res.get("wieder_da_kinder") or []):
                hit = _ki_kid(kn)
                if hit:
                    g, disp = hit
                    if disp in fa.get(g, []):
                        fa[g].remove(disp)
                    note_lines.append(f"{disp} wieder da")
            for tn in (res.get("abwesend_trainer") or []):
                d = normalize_name((tn or "").strip())
                if d in ALLE_TRAINER:
                    fa.setdefault("Trainer", [])
                    if d not in fa["Trainer"]:
                        fa["Trainer"].append(d)
                    note_lines.append(f"{d} nicht da")
            eint = res.get("einteilung") or {}
            if isinstance(eint, dict) and eint:
                val = []
                for tr, lab in eint.items():
                    d = normalize_name((tr or "").strip())
                    if d in ALLE_TRAINER and lab:
                        val.append({"trainer": d, "gruppe": str(lab)})
                if val:
                    ki["assign"] = _merge_ki_assign(ki.get("assign"), val)
            # Umgesetzte Anweisungen werden NICHT als Anmerkungs-Text in den Plan geschrieben
            # (Noah 16.07.2026). Nur echter Hinweistext ("In die Anmerkungen schreiben: ...")
            # landet als Notiz; note_lines dienen nur noch dem Log unten.
            base_note = (res.get("notiz") or "").strip()
            if base_note:
                prev = (fe.get("notiz") or "").strip()
                fe["notiz"] = (prev + "\n" if prev else "") + base_note
            if a.get("id"):
                processed_ids.append(a["id"])
            if a in anmerkungen_server:
                anmerkungen_server.remove(a)
            changed = True
            print(f"[KI] '{notiz[:45]}' ({writer}) -> {dk}")
        except Exception as e:
            print(f"[KI] Anwenden fehlgeschlagen: {e}")
    if changed:
        try:
            f = sftp.open("fixed_entries.json", "wb")
            f.write(json.dumps(fixed_entries, indent=2, ensure_ascii=False).encode("utf-8"))
            f.close()
            print("[KI] fixed_entries.json aktualisiert.")
        except Exception as e:
            print(f"[KI] fixed_entries Upload fehlgeschlagen: {e}")
    if processed_ids:
        try:
            mark_anmerkungen_gelesen(sftp, [i for i in processed_ids if i])
        except Exception as e:
            print(f"[KI] mark gelesen fehlgeschlagen: {e}")


def main():
    print(f"=== TV Rheinzabern Auto-Trainingsplan | {date.today()} ===\n")

    print("Verbinde mit Server...")
    ssh, sftp = get_sftp()
    apply_config_roster(sftp)
    require_server_roster()

    state              = load_state(sftp)
    abmeldungen, raw_abm_hash = read_abmeldungen(sftp)
    anmerkungen_server = read_anmerkungen_server(sftp)
    fixed_entries      = read_fixed_entries(sftp)
    print(f"Abmeldungen geladen: {len(abmeldungen)} Eintraege")
    print(f"Ungelesene Anmerkungen: {len(anmerkungen_server)}")
    print(f"Abmeldungen-Hash (raw): {raw_abm_hash[:8]}...")

    today         = date.today()
    training_date = active_training_date()
    datum         = fmt_datum(training_date)
    datum_kurz    = fmt_datum_kurz(training_date)
    wtag          = wochentag_name(training_date)
    days_away     = (training_date - today).days

    print(f"\nNaechstes Training: {wtag} {datum} ({days_away} Tage)\n")

    # Abwesenheiten und Hash berechnen
    tag = tag_of_date(training_date)
    grid_rows, grid_phase = compute_time_grid(GRUPPEN_ZEITEN, tag)
    absences, late_notes, trainer_timing = get_absences(abmeldungen, training_date, grid_rows)
    late_notes = late_notes + timing_annotations(trainer_timing)
    # raw_abm_hash für konsistenten Vergleich mit check_quick.py (beide nutzen raw JSON hash)
    new_hash = raw_abm_hash

    # -- Deterministischer Kommando-Parser (Bug-Fix 19.08.2026): erkennt einfache
    # Kurz-Kommandos wie "Noah Springer" oder "lösche alle Anmerkungen" OHNE KI
    # und wendet sie direkt an. Was hier nicht matcht, bleibt fuer die KI. --
    try:
        preprocess_deterministic_annotations(sftp, anmerkungen_server, fixed_entries)
    except Exception as _cmde:
        print(f"[CMD] uebersprungen: {_cmde}")

    # -- Mini-KI: freie Trainer-Anmerkungen analysieren und in fixed_entries anwenden --
    try:
        ki_process_annotations(sftp, anmerkungen_server, fixed_entries, state)
    except Exception as _kie:
        print(f"[KI] uebersprungen: {_kie}")

    # Fixed entries anwenden (erzwingt Abwesenheiten, die das Auto-System nicht ändern darf)
    fixed_for_date = fixed_entries.get(datum_kurz, {})
    lock_trainer   = fixed_for_date.get("lock_trainer_plan", False)
    fixed_tp_raw   = fixed_for_date.get("fixed_trainer_plan")  # vorberechneter Trainer-Plan
    if fixed_for_date:
        absences = apply_fixed_absences(absences, fixed_for_date)
        if lock_trainer:
            print(f"[FIXED] Trainer-Plan für {datum_kurz} ist gesperrt – Auto-Berechnung deaktiviert.")

    # ── Trainingsentfall-Check (Training wurde auf der Website abgesagt) ──────────
    # trainingsentfall.json ist die Single Source of Truth. Ist das nächste Training
    # als Entfall markiert, wird KEIN normaler Plan erzeugt, sondern der Entfall-Hinweis
    # veröffentlicht. Wird der Entfall wieder aufgehoben, erzeugt das System einen frischen Plan.
    entfall_list      = read_trainingsentfall(sftp)
    datum_iso         = training_date.strftime("%Y-%m-%d")
    entfall_published = state.setdefault("entfall_published", [])

    # Bugfix 30.08.2026 (Noah: Training vom 26.08. war ausgefallen, stand aber nicht
    # als Entfall im Trainingsplan): Dieser Check pruefte bisher AUSSCHLIESSLICH das
    # naechste anstehende Training (datum_iso == active_training_date()). Jeder andere
    # Eintrag in trainingsentfall.json -- insbesondere ein bereits vergangenes Training,
    # das erst nachtraeglich als Entfall markiert wurde -- wurde dadurch nie
    # veroeffentlicht, selbst wenn er weiterhin in der Liste stand. Jetzt werden zuerst
    # ALLE anderen Eintraege der Liste nachgeholt, bevor wie gehabt das naechste
    # Training geprueft wird.
    # Hotfix 30.08.2026: trainingsentfall.json sammelt seit Monaten Karteileichen
    # (laengst vergangene, nie aus der Liste entfernte Entfall-Daten). Der erste
    # Lauf mit dem Fix oben hat das schmerzhaft gezeigt: 15.05./12.06.2026 waren
    # nie in entfall_published vermerkt und loesten dadurch eine ECHTE, Monate zu
    # spaete WhatsApp-/E-Mail-Benachrichtigung aus. Ab jetzt werden nur Eintraege
    # aus den letzten 14 Tagen (der eigentliche Anwendungsfall: ein kuerzlich
    # nachtraeglich markiertes Training wie der 26.08.-Fall) automatisch nachgeholt.
    # Aeltere Karteileichen werden weder angefasst noch benachrichtigt.
    _other_published = False
    for _ef_iso in entfall_list:
        if _ef_iso == datum_iso or not _entfall_is_recent(_ef_iso, today):
            continue
        if _publish_entfall_for(sftp, state, fixed_entries, _ef_iso, entfall_published):
            _other_published = True
    if _other_published:
        save_state(sftp, state)

    if datum_iso in entfall_list:
        _published_now = _publish_entfall_for(sftp, state, fixed_entries, datum_iso, entfall_published)
        save_state(sftp, state)
        if not _published_now:
            print(f"[ENTFALL] {datum} bereits als Entfall veröffentlicht – nichts zu tun.")
        sftp.close(); ssh.close()
        return
    elif datum_kurz in entfall_published:
        # Entfall wurde wieder aufgehoben → Hinweis entfernen, frischen Plan erzeugen
        print(f"[ENTFALL] Entfall für {datum} aufgehoben → normaler Plan wird neu erstellt.")
        entfall_published.remove(datum_kurz)
        state.setdefault("entfall_notiz_hash", {}).pop(datum_kurz, None)
        remove_plan_files(sftp, datum_kurz)
        state.get("plan_data", {}).pop(datum_kurz, None)
        if datum_kurz in state.get("generated_plans", []):
            state["generated_plans"].remove(datum_kurz)
        save_state(sftp, state)

    # -- KI-Timing (aus Anmerkung) in trainer_timing mergen -> in allen Pfaden rot geblockt --
    try:
        for _kn, _kv in ki_timing_to_dict((fixed_for_date.get("ki") or {}).get("timing"), grid_rows).items():
            trainer_timing[_kn] = _kv
    except Exception:
        pass

    # -- Admin-Editor (manuell_bearbeitet aus admin.php) anwenden --------------
    force_regen = False
    admin_fixed_hash = ""
    if fixed_for_date.get("manuell_bearbeitet"):
        import hashlib as _hl
        admin_fixed_hash = _hl.md5(json.dumps(fixed_for_date, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()
        _ki = fixed_for_date.get("ki") or {}
        _partial = fixed_for_date.get("fixed_trainer_partial") or {}
        _troles_hist = _load_trainer_roles_history(state, exclude_date=datum_kurz)
        if _partial:
            # Bugfix 19.08.2026 (Noah): Admin-Zellen sind eine harte Vorgabe, die KI/
            # Kommando-Anweisungen (ki.assign/merges) planen - falls vorhanden - den
            # Rest drumherum, statt komplett ignoriert zu werden (siehe build_admin_trainer_plan).
            _base_tp = build_admin_trainer_plan(absences, _partial, grid_rows, grid_phase, trainer_roles_history=_troles_hist, ki=_ki)
        elif _ki.get("assign") or _ki.get("merges"):
            try:
                _base_tp, _sd, _ka = build_ki_einteilung(absences, _ki, grid_rows, grid_phase, _troles_hist)
            except Exception as _kierr:
                print(f"[KI] Einteilung fehlgeschlagen, Fallback build_admin: {_kierr}")
                _base_tp = build_admin_trainer_plan(absences, _partial, grid_rows, grid_phase, trainer_roles_history=_troles_hist, ki=_ki)
        else:
            _base_tp = build_admin_trainer_plan(absences, _partial, grid_rows, grid_phase, trainer_roles_history=_troles_hist, ki=_ki)
        fixed_for_date["fixed_trainer_plan"] = _base_tp
        fixed_for_date["lock_trainer_plan"] = True
        lock_trainer = True
        fixed_tp_raw = _base_tp
        _notiz = (fixed_for_date.get("notiz") or "").strip()
        if _notiz:
            anmerkungen_server.append({"trainer": "Hinweis", "notiz": _notiz})
        _stored_fh = state.get("plan_data", {}).get(datum_kurz, {}).get("fixed_hash", "")
        if admin_fixed_hash != _stored_fh:
            print(f"[ADMIN] Manuelle Planbearbeitung {datum_kurz} (neu/geaendert) -> regeneriere.")
            remove_plan_files(sftp, datum_kurz)
            state.get("plan_data", {}).pop(datum_kurz, None)
            if datum_kurz in state.get("generated_plans", []):
                state["generated_plans"].remove(datum_kurz)
            force_regen = True

    if plan_exists(sftp, datum_kurz):
        # ── Geschützter Plan → nie überschreiben ───────────────
        if datum_kurz in state.get("protected_plans", []):
            print(f"Plan {datum_kurz} ist als geschützt markiert – wird nicht überschrieben.")
            sftp.close(); ssh.close()
            return

        # ── Plan vorhanden: prüfe ob Update nötig ──────────────
        plan_data   = state.get("plan_data", {}).get(datum_kurz, {})
        stored_hash = plan_data.get("absences_hash", "")

        if not plan_data or not stored_hash:
            # Plan ohne gespeicherten Hash (manuell erstellt oder erste Initialisierung)
            # → Plan mit AKTUELLEN Abwesenheiten NEU ERSTELLEN (nicht nur Hash speichern)
            print(f"Plan ohne Hash fuer {datum_kurz} → Plan wird mit aktuellen Abwesenheiten erstellt.")
            _ng_idx = get_next_gear(state, exclude_date=datum_kurz, fixed_entries=fixed_entries)
            geraet_combo = GERAETE_ROTATION[_ng_idx]
            # plan_data mit Defaults initialisieren und stored_hash="" setzen
            # damit die Update-Logik unten sicher ausgeführt wird
            plan_data = {
                "absences_hash":    "",   # leer → abs_changed=True → Plan wird generiert
                "trainer_absences": list(absences.get("Trainer", [])),
                "stored_absences":  {},
                "geraet_1":         geraet_combo[0],
                "geraet_2":         geraet_combo[1],
            }
            state.setdefault("plan_data", {})[datum_kurz] = plan_data
            if datum_kurz not in state.get("generated_plans", []):
                state.setdefault("generated_plans", []).append(datum_kurz)
            stored_hash = ""   # explizit leer → Update-Logik greift

        has_new_anm = len(anmerkungen_server) > 0
        abs_changed = (new_hash != stored_hash)

        if not abs_changed and not has_new_anm and not FORCE_REGEN:
            print("Plan aktuell, keine Aenderungen → nichts zu tun.")
            sftp.close(); ssh.close()
            return

        if abs_changed:
            print(f"Abmeldungsaenderung erkannt (Hash {stored_hash[:8]}... → {new_hash[:8]}...)")
        if has_new_anm:
            print(f"Neue Trainer-Anmerkungen: {len(anmerkungen_server)}")
        if FORCE_REGEN and not abs_changed and not has_new_anm:
            print("[FORCE] FORCE_REGEN gesetzt -> Regenerierung ohne Datenaenderung erzwungen.")

        # Trainer-Absences prüfen
        stored_trainer_abs = set(plan_data.get("trainer_absences", []))
        new_trainer_abs    = set(absences.get("Trainer", []))
        fixed_trainer_abs  = set(fixed_for_date.get("fixed_absences", {}).get("Trainer", []))

        # Trainer-Änderung ignorieren wenn der Plan gesperrt ist (lock_trainer_plan)
        trainer_changed = (stored_trainer_abs != new_trainer_abs)
        if trainer_changed and lock_trainer:
            print(f"[FIXED] Trainer-Änderung ignoriert (Plan gesperrt): "
                  f"{stored_trainer_abs} → {new_trainer_abs}")
            trainer_changed = False

        # Kurze Hinweise sammeln -> am Ende EINE konsolidierte Mail statt mehrerer
        hinweise = []

        if trainer_changed:
            added   = new_trainer_abs - stored_trainer_abs
            removed = stored_trainer_abs - new_trainer_abs
            wer = ', '.join(sorted(added | removed)) or '–'
            hinweise.append(f"die Trainereinteilung wurde geändert, da {wer} abgemeldet ist")

        # UPDATE: Plan wird IMMER automatisch (neu) erstellt - auch bei Trainerwechsel
        # oder Engpass. Noah muss dafuer nichts mehr manuell machen (07.07.2026).
        print("UPDATE: erstelle Plan automatisch neu...")

        geraet_1     = plan_data["geraet_1"]
        geraet_2     = plan_data["geraet_2"]

        issues, anwesend_trainer, soft_warnings = detect_complex(absences, late_notes)
        # Bei gesperrtem Trainer-Plan: Trainer-Anzahl-Fehler ignorieren
        if lock_trainer:
            issues = [i for i in issues if "Trainer anwesend" not in i]
        if issues:
            # Auch bei Problemen: trotzdem automatisch fertigstellen, nur kurz vermerken
            hinweise.append("beim automatischen Erstellen gab es ein Problem, bitte kurz prüfen")
            print(f"[WARN] Issues trotzdem automatisch weitergebaut: {issues}")

        if lock_trainer and fixed_tp_raw:
            # Gesperrter Trainer-Plan: direkt aus fixed_entries (UPDATE path)
            trainer_plan = {
                k: [tuple(s) for s in v] if v is not None else None
                for k, v in fixed_tp_raw.items()
            }
            sondertiming = {}
            anmerkungen  = []
            trainer_plan = apply_timing_coverage(trainer_plan, trainer_timing)
            trainer_plan = apply_timing_blocks(trainer_plan, trainer_timing)
            print("[FIXED] Verwende gesperrten Trainer-Plan aus fixed_entries (UPDATE).")
        else:
            _troles_hist = _load_trainer_roles_history(state, exclude_date=datum_kurz)
            trainer_plan, sondertiming, anmerkungen = build_trainer_plan(
                absences, grid_rows, grid_phase,
                trainer_roles_history=_troles_hist,
            )
            trainer_plan = apply_timing_coverage(trainer_plan, trainer_timing)
            trainer_plan = apply_timing_blocks(trainer_plan, trainer_timing)

        # Verspätungen / frühes Gehen als Hinweis eintragen
        for note in late_notes:
            anmerkungen.append(note)

        # Trainer-Anmerkungen eintragen; Admin-"Hinweis" separat als Override
        _admin_notiz = ""
        for anm in anmerkungen_server:
            trainer_name = anm.get("trainer", "")
            notiz        = anm.get("notiz", "").strip()
            if not notiz:
                continue
            if trainer_name == "Hinweis":
                _admin_notiz = notiz
                continue
            anmerkungen.append(f"• {trainer_name}: {notiz}")

        # Auto-Anmerkungen fuer Admin-Vorbefuellung sichern (vor evtl. Override)
        upload_anmerkungen_auto(sftp, datum_kurz, anmerkungen)

        # Admin-Override: editierte Notiz ersetzt die Anmerkungen 1:1 (kein Doppeln)
        if _admin_notiz:
            anmerkungen = [ln.rstrip() for ln in _admin_notiz.split("\n") if ln.strip()]

        xlsx_path, pdf_path = build_excel(
            datum=datum, wochentag=wtag,
            geraet_1=geraet_1, geraet_2=geraet_2,
            abwesend=absences,
            trainer_plan=trainer_plan,
            sondertiming=sondertiming,
            anmerkungen=anmerkungen,
            grid_rows=grid_rows,
        )
        upload_pdf(sftp, pdf_path, datum_kurz)
        upload_xlsx(sftp, xlsx_path, datum_kurz)
        aktuell_json = build_aktuell_json(datum, datum_kurz, wtag, geraet_1, geraet_2, trainer_plan, absences, grid_rows)
        upload_aktuell_json(sftp, aktuell_json)

        ids_gelesen = [a["id"] for a in anmerkungen_server if a.get("id")]
        mark_anmerkungen_gelesen(sftp, ids_gelesen)

        # Dedup: late_notes
        already_sent_notes    = set(plan_data.get("late_notes_sent", []))
        new_late_notes        = [n for n in late_notes if n not in already_sent_notes]

        # Dedup: soft warnings (≤2 Turner in Gruppe, Engpass, ...)
        already_sent_warnings = set(plan_data.get("warnings_sent", []))
        new_soft_warnings     = [(k, t) for k, t in soft_warnings if k not in already_sent_warnings]

        # Dedup: Trainer-Anmerkungen (nur neue IDs sollen einen Hinweis ausloesen)
        already_sent_anm_ids  = set(plan_data.get("anm_notified_ids", []))
        new_anm_ids           = [a["id"] for a in anmerkungen_server if a.get("id") and a["id"] not in already_sent_anm_ids]

        # State: Hash + dedup-Listen aktualisieren (auch trainer_absences - wichtig,
        # da dieser Block jetzt auch bei Trainerwechsel laeuft)
        plan_data["absences_hash"]    = new_hash
        plan_data["trainer_absences"] = list(absences.get("Trainer", []))
        # Trainer-Rollen fuer Rotation im naechsten Plan speichern (Bug-Fix 19.08.2026)
        plan_data["trainer_roles"]    = _extract_trainer_roles(trainer_plan)
        if admin_fixed_hash:
            plan_data["fixed_hash"] = admin_fixed_hash
        plan_data["stored_absences"] = absences
        if new_late_notes:
            plan_data["late_notes_sent"] = list(already_sent_notes | set(new_late_notes))
        if new_soft_warnings:
            plan_data["warnings_sent"] = list(already_sent_warnings | {k for k, _ in new_soft_warnings})
        if new_anm_ids:
            plan_data["anm_notified_ids"] = list(already_sent_anm_ids | set(new_anm_ids))
        save_state(sftp, state)

        # Weitere kurze Hinweise sammeln - NUR was wirklich eine Aktion/Aenderung war
        # (Trainerwechsel, Engpass-Zusammenlegung, Trainer-Anmerkung, Verspaetung).
        # Reine Turner-bezogene Warnungen (z.B. "nur 2 Turner in Gruppe X") ohne
        # Handlungsbedarf loesen bewusst KEINE Mail aus - nur Log (Noah, 07.07.2026).
        if new_anm_ids:
            hinweise.append("es gibt neue Trainer-Anmerkungen")
        if new_late_notes:
            hinweise.append("es gibt neue Verspätungs-/Frühgeh-Hinweise")
        low_trainer_new = [(k, t) for k, t in new_soft_warnings if k.startswith("low_trainer_")]
        if low_trainer_new:
            hinweise.append("nur wenige Trainer da, Gruppen wurden zusammengelegt")
        other_new_warnings = [t for k, t in new_soft_warnings if not k.startswith("low_trainer_")]
        if other_new_warnings:
            print(f"[INFO] Reine Turner-Hinweise fuer {datum} (keine Mail, kein Handlungsbedarf): {other_new_warnings}")

        # EINE konsolidierte, kurze Mail - nur wenn es etwas zu berichten gibt.
        # Reine Turner-Abwesenheitsaenderungen ohne weitere Auffaelligkeit -> keine Mail.
        if hinweise:
            send_whatsapp(
                f"Hallo, der Trainingsplan für {datum} wurde automatisch aktualisiert: "
                + "; ".join(hinweise) + ". Bitte bei Gelegenheit kurz prüfen."
            )
            print(f"UPDATE FERTIG fuer {datum}. Mail gesendet: {hinweise}")
        else:
            print(f"UPDATE FERTIG fuer {datum}. Keine Mail (nur Turner-Abwesenheiten geändert).")

    else:
        # ── Kein Plan vorhanden: erstelle neuen ────────────────
        # NUR im Publikationsfenster (Mi oder Fr nach 22:00 CEST / 20:00 UTC) erstellen
        now_utc   = datetime.now(timezone.utc)
        days_away = (training_date - date.today()).days

        # Sicherheitsnetz: Steht das naechste Training unmittelbar bevor (<=1 Tag)
        # und existiert noch KEIN Plan, wird sofort erzeugt - auch ausserhalb des
        # 22:00-Fensters. So bleibt kein Training ohne Plan, falls das regulaere
        # Mi/Fr-Abendfenster verpasst wurde.
        imminent = days_away <= 1
        if not is_publication_window() and not force_regen and not imminent:
            print(
                f"Kein Plan vorhanden, aber außerhalb Publikationsfenster "
                f"({now_utc.strftime('%H:%M')} UTC, Wochentag {now_utc.weekday()}) → warte bis Mi/Fr 22:00 CEST."
            )
            sftp.close(); ssh.close()
            return
        if imminent and not is_publication_window() and not force_regen:
            print(f"[SICHERHEITSNETZ] Training in {days_away} Tag(en), kein Plan → erstelle sofort (Fenster uebergangen).")

        if days_away > 5 and not force_regen:
            print(
                f"Kein Plan vorhanden, Training in {days_away} Tagen "
                f"({now_utc.strftime('%H:%M')} UTC) → noch zu weit entfernt."
            )
            sftp.close(); ssh.close()
            return

        print(f"Kein Plan vorhanden, starte Generierung... (Training in {days_away} Tagen, {now_utc.strftime('%H:%M')} UTC)")

        # Geräte-Rotation: letzte (bis zu) 8 ECHTEN Trainings ansehen, ausgefallene
        # ignorieren und genau einen Schritt weiterrücken (robust gegen Drift).
        new_combo_idx = get_next_gear(state, exclude_date=datum_kurz, fixed_entries=fixed_entries)
        geraet_1, geraet_2 = GERAETE_ROTATION[new_combo_idx]

        # Bei gesperrtem Trainer-Plan: Geräte aus fixed_entries übernehmen (falls vorhanden)
        if lock_trainer and fixed_for_date.get("geraet_1"):
            geraet_1     = fixed_for_date["geraet_1"]
            geraet_2     = fixed_for_date["geraet_2"]
            print(f"[FIXED] Geräte aus fixed_entries: {geraet_1} + {geraet_2}")

        # Kurze Hinweise sammeln -> am Ende EINE konsolidierte Mail statt mehrerer
        hinweise = []

        issues, anwesend_trainer, soft_warnings = detect_complex(absences, late_notes)
        # Bei gesperrtem Trainer-Plan: Trainer-Anzahl-Fehler ignorieren (NEW path)
        if lock_trainer:
            issues = [i for i in issues if "Trainer anwesend" not in i]

        if issues:
            # Auch bei Problemen: trotzdem automatisch fertigstellen, nur kurz vermerken
            hinweise.append("beim automatischen Erstellen gab es ein Problem, bitte kurz prüfen")
            print(f"[WARN] Issues trotzdem automatisch weitergebaut: {issues}")

        if lock_trainer and fixed_tp_raw:
            # Gesperrter Trainer-Plan: direkt aus fixed_entries (NEW path)
            trainer_plan = {
                k: [tuple(s) for s in v] if v is not None else None
                for k, v in fixed_tp_raw.items()
            }
            sondertiming = {}
            anmerkungen  = []
            trainer_plan = apply_timing_coverage(trainer_plan, trainer_timing)
            trainer_plan = apply_timing_blocks(trainer_plan, trainer_timing)
            print("[FIXED] Verwende gesperrten Trainer-Plan aus fixed_entries (NEW).")
        else:
            _troles_hist = _load_trainer_roles_history(state, exclude_date=datum_kurz)
            trainer_plan, sondertiming, anmerkungen = build_trainer_plan(
                absences, grid_rows, grid_phase,
                trainer_roles_history=_troles_hist,
            )
            trainer_plan = apply_timing_coverage(trainer_plan, trainer_timing)
            trainer_plan = apply_timing_blocks(trainer_plan, trainer_timing)

        # Verspätungen / frühes Gehen als Hinweis eintragen
        for note in late_notes:
            anmerkungen.append(note)

        # Trainer-Anmerkungen eintragen; Admin-"Hinweis" separat als Override
        _admin_notiz = ""
        for anm in anmerkungen_server:
            trainer_name = anm.get("trainer", "")
            notiz        = anm.get("notiz", "").strip()
            if not notiz:
                continue
            if trainer_name == "Hinweis":
                _admin_notiz = notiz
                continue
            anmerkungen.append(f"• {trainer_name}: {notiz}")

        # Auto-Anmerkungen fuer Admin-Vorbefuellung sichern (vor evtl. Override)
        upload_anmerkungen_auto(sftp, datum_kurz, anmerkungen)

        # Admin-Override: editierte Notiz ersetzt die Anmerkungen 1:1 (kein Doppeln)
        if _admin_notiz:
            anmerkungen = [ln.rstrip() for ln in _admin_notiz.split("\n") if ln.strip()]

        xlsx_path, pdf_path = build_excel(
            datum=datum,
            wochentag=wtag,
            geraet_1=geraet_1,
            geraet_2=geraet_2,
            abwesend=absences,
            trainer_plan=trainer_plan,
            sondertiming=sondertiming,
            anmerkungen=anmerkungen,
            grid_rows=grid_rows,
        )

        upload_pdf(sftp, pdf_path, datum_kurz)
        upload_xlsx(sftp, xlsx_path, datum_kurz)
        aktuell_json = build_aktuell_json(datum, datum_kurz, wtag, geraet_1, geraet_2, trainer_plan, absences, grid_rows)
        upload_aktuell_json(sftp, aktuell_json)

        ids_gelesen = [a["id"] for a in anmerkungen_server if a.get("id")]
        mark_anmerkungen_gelesen(sftp, ids_gelesen)

        # State aktualisieren
        state["last_training_date"] = datum
        state["geraet_combo_index"] = new_combo_idx
        state.setdefault("generated_plans", []).append(datum_kurz)
        state.setdefault("plan_data", {})[datum_kurz] = {
            "absences_hash":    new_hash,
            "trainer_absences": list(absences.get("Trainer", [])),
            "trainer_roles":    _extract_trainer_roles(trainer_plan),  # fuer Rotation im naechsten Plan
            "stored_absences":  absences,
            "geraet_1":         geraet_1,
            "geraet_2":         geraet_2,
            "late_notes_sent":  list(late_notes),            # bereits gesendet beim Erstellen
            "warnings_sent":    [k for k, _ in soft_warnings],  # bereits gesendet beim Erstellen
            "anm_notified_ids": [a["id"] for a in anmerkungen_server if a.get("id")],  # bereits gemeldet
            "fixed_hash":       admin_fixed_hash,
        }
        save_state(sftp, state)

        # Kurze Hinweise sammeln - fliessen in die EINE Erstellungs-Mail mit ein.
        # Reine Turner-bezogene Warnungen ohne Handlungsbedarf -> nur Log, keine Erwaehnung.
        if anmerkungen_server:
            hinweise.append("es gibt neue Trainer-Anmerkungen")
        if late_notes:
            hinweise.append("es gibt Verspätungs-/Frühgeh-Hinweise")
        low_trainer_warn = [(k, t) for k, t in soft_warnings if k.startswith("low_trainer_")]
        if low_trainer_warn:
            hinweise.append("nur wenige Trainer da, Gruppen wurden zusammengelegt")
        other_warnings = [t for k, t in soft_warnings if not k.startswith("low_trainer_")]
        if other_warnings:
            print(f"[INFO] Reine Turner-Hinweise fuer {datum} (keine Erwaehnung, kein Handlungsbedarf): {other_warnings}")

        # Haupt-Notification: Plan fertig (immer genau EINE Mail, kurz, mit Hinweisen falls vorhanden)
        text = f"Hallo, der Trainingsplan für {wtag}, {datum} ist erstellt."
        if hinweise:
            text += " " + "; ".join(hinweise).capitalize() + ". Bitte bei Gelegenheit kurz prüfen."
        send_whatsapp(text)
        print(f"FERTIG! Plan fuer {datum} hochgeladen. Hinweise: {hinweise or 'keine'}")

    sftp.close()
    ssh.close()

if __name__ == "__main__":
    main()
# (Trainingsentfall-Support 19.06.2026)
