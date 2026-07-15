#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
create_wochenchallenge_v2.py
============================
Reusable Wochenchallenge-Generator fuer TV Rheinzabern.
Claude passt NUR den CONFIG-Block an und fuehrt das Script aus.
Abwesenheiten werden automatisch aus abmeldungen.json geladen.

Ausfuehrung (Sandbox):
    python3 /sessions/.../mnt/Claude/create_wochenchallenge_v2.py
"""

import os, json
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ===================================================================
#  CONFIG - NUR DIESEN BLOCK ANPASSEN
# ===================================================================

START_DATUM = "06.06.2026"   # Samstag der Woche (TT.MM.JJJJ)
END_DATUM   = "09.06.2026"   # Dienstag der Woche

# Spaltenbeschriftungen
TAGE_HEADER = ["Sa\n06.06.", "So\n07.06.", "Mo\n08.06.", "Di\n09.06."]

# Trainings: Name -> Liste der Tage (0=Sa, 1=So, 2=Mo, 3=Di)
GRUPPEN = {
    "G1": {
        "Felix":    [0, 1, 2, 3],
        "Finn":     [0, 1, 2],
        "Sinan":    [],
        "Ilyas":    [0, 1, 2, 3],
        "Jonathan": [1, 2, 3],
        "Hannes":   [1, 3],
        "Ben G1":   [0, 1, 2, 3],
    },
    "G2": {
        "Henry":  [0, 1, 2, 3],
        "Matti":  [0, 1, 2, 3],
        "Levent": [0, 1, 2, 3],
        "Caius":  [0, 1, 2, 3],
    },
    "G3": {
        "Artem":   [1, 2],
        "Finn":    [0, 1, 2, 3],
        "Ben G3":  [2],
        "Erik":    [0, 1, 3],
        "Michael": [0, 1, 3],
    },
    "G4": {
        "Felix": [],
        "Mika":  [0, 2, 3],
        "Anton": [],
        "Jamie": [],
    },
}

# Vorherige Gesamtpunkte (Stand vor dieser Woche)
VORHERIGE_PUNKTE = {
    "G1_Felix": 4, "G1_Finn": 0, "G1_Sinan": 0,
    "G1_Ilyas": 4, "G1_Jonathan": 4, "G1_Hannes": 3, "G1_Ben G1": 0,
    "G2_Henry": 4, "G2_Matti": 4, "G2_Levent": 4, "G2_Caius": 4,
    "G3_Artem": 0, "G3_Finn": 0, "G3_Ben G3": 1, "G3_Erik": 0, "G3_Michael": 0,
    "G4_Felix": 3, "G4_Mika": 18, "G4_Anton": 3, "G4_Jamie": 0,
}

# ===================================================================
#  FARBEN & DESIGN (nicht aendern)
# ===================================================================

C_TITLE_BG    = "1A1A2E"
C_SUBTITLE_BG = "16213E"
C_HEADER_TAGE = "0F3460"
C_HEADER_ERG  = "00B4D8"
C_HEADER_INSG = "E67E22"
C_WHITE       = "FFFFFF"
C_ROW_LIGHT   = "E8F4FD"
C_ROW_GREY    = "F5F5F5"
C_SEPARATOR   = "D0E8F5"
C_LEGEND_BG   = "EEF6FF"
C_CHECK_BG    = "C8E6C9"
C_CHECK_FG    = "1B5E20"
C_RED         = "FF4444"
C_ORANGE      = "FF8C00"
C_GREEN       = "2ECC71"
C_INSG_BG     = "FFF3E0"

# ===================================================================
#  AUTO-ABWESENHEITEN (nicht aendern)
#  Laedt abmeldungen.json, filtert auf END_DATUM+1 (=Mittwoch nach der WC)
# ===================================================================

def _load_abwesend():
    """
    Lädt Abwesenheiten für den Mittwoch nach der WC (END_DATUM + 1 Tag).
    Sucht abmeldungen.json zuerst im Skript-Verzeichnis, dann unter /tmp/.
    Normalisiert Namen aus altem Format (Sinan → Sinan Y.) auf neues Format.
    """
    # Name-Normalisierung: altes Format → neues Format (Vorname N.)
    NAME_NORMALIZE = {
        "Sinan":        "Sinan Y.",   "Ilyas":       "Ilyas E.",   "Jonathan":    "Jonathan S.",
        "Hannes":       "Hannes G.",  "Henry":       "Henry K.",   "Matti":       "Matti G.",
        "Levent":       "Levent K.",  "Caius":       "Caius C.",   "Erik":        "Erik E.",
        "Artem":        "Artem T.",   "Michael":     "Michael K.", "Anton":       "Anton K.",
        "Mika":         "Mika W.",    "Jamie":       "Jamie G.",
        "Felix (G1)":   "Felix E.",   "Finn (G1)":   "Finn M.",    "Ben (G1)":    "Ben B.",
        "Ben G1":       "Ben B.",     "Felix G1":    "Felix E.",   "Finn G1":     "Finn M.",
        "Finn (G3)":    "Finn T.",    "Ben (G3)":    "Ben F.",     "Finn G3":     "Finn T.",
        "Ben G3":       "Ben F.",     "Felix (G4)":  "Felix L.",   "Felix G4":    "Felix L.",
    }

    end = datetime.strptime(END_DATUM, "%d.%m.%Y")
    check_date = (end + timedelta(days=1)).strftime("%Y-%m-%d")

    # Suche abmeldungen.json: erst Skript-Verzeichnis, dann /tmp/
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(script_dir, "abmeldungen.json"),
        "/tmp/abmeldungen.json",
    ]
    json_path = None
    for p in candidates:
        if os.path.exists(p):
            json_path = p
            break

    if not json_path:
        print(f"[ABWESENHEIT] abmeldungen.json nicht gefunden – keine Abwesenheitsmarkierung")
        return []

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    abwesend = []
    for entry in data:
        if entry.get("datum") == check_date and entry.get("gruppe", "").startswith("G"):
            raw = entry.get("name", "").strip()
            # Normalisieren: erst Mapping prüfen, dann Klammern entfernen als Fallback
            name = NAME_NORMALIZE.get(raw, raw)
            if "(" in name:
                name = name.split("(")[0].strip()
            if name and name not in abwesend:
                abwesend.append(name)

    print(f"[ABWESENHEIT] {check_date} (aus {os.path.basename(json_path)}): {abwesend if abwesend else 'keine'}")
    return abwesend

ABWESEND = _load_abwesend()

# ===================================================================
#  PFADE (automatisch)
# ===================================================================

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
datum_short = f"{START_DATUM[0:2]}.{START_DATUM[3:5]}.{START_DATUM[8:10]}"
ORDNER      = os.path.join(SCRIPT_DIR, "Wochenchallenge", f"ab {datum_short}")
XLSX_PATH   = os.path.join(ORDNER, f"ab_{datum_short}_Wochenchallenge.xlsx")
PDF_PATH    = os.path.join(ORDNER, f"ab_{datum_short}_Wochenchallenge.pdf")
os.makedirs(ORDNER, exist_ok=True)

# ===================================================================
#  HILFSFUNKTIONEN
# ===================================================================

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(size=11, bold=False, color="000000", italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)
def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def medium():
    s = Side(style="medium", color="1A1A2E")
    return Border(left=s, right=s, top=s, bottom=s)
def get_vor(gruppe, name):
    return VORHERIGE_PUNKTE.get(f"{gruppe}_{name}", 0)

def _mp_state():
    import json as _j, os as _o
    _p=_o.path.join(_o.path.dirname(_o.path.abspath(__file__)),"murmel_punkte_state.json")
    if not _o.path.exists(_p): return {"evalBase":{},"redeemed":{}}
    try: _d=_j.load(open(_p,encoding="utf-8"))
    except Exception: return {"evalBase":{},"redeemed":{}}
    _ev=_d.get("evalBase"); _rd=_d.get("redeemed")
    return {"evalBase":_ev if isinstance(_ev,dict) else {}, "redeemed":_rd if isinstance(_rd,dict) else {}}
_MP_STATE=_mp_state()
def get_reset(gruppe, name):
    """Board-Reset-Stand: G1/G2 -> evalBase, G3/G4 -> redeemed. Namen bereits kanonisch."""
    if gruppe in ("G1","G2"): return int(_MP_STATE["evalBase"].get(name,0))
    return int(_MP_STATE["redeemed"].get(name,0))

# ===================================================================
#  EXCEL ERSTELLEN
# ===================================================================

def create_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wochenchallenge"

    ws.column_dimensions["A"].width = 22
    for col in ["B","C","D","E"]: ws.column_dimensions[col].width = 13
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 17

    # Titel
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    s_d = f"{START_DATUM[0:2]}.{START_DATUM[3:5]}."
    e_d = f"{END_DATUM[0:2]}.{END_DATUM[3:5]}.{END_DATUM[6:10]}"
    c.value     = f"\U0001F3C6 WOCHENCHALLENGE – {s_d} – {e_d}"
    c.fill      = fill(C_TITLE_BG)
    c.font      = fnt(16, bold=True, color=C_WHITE)
    c.alignment = aln("center","center")

    # Untertitel
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value     = "1 Trainingstag = 1 Punkt | Max. 4 Punkte | 0–1 \U0001F534  2 \U0001F7E0  3–4 \U0001F7E2"
    c.fill      = fill(C_SUBTITLE_BG)
    c.font      = fnt(10, italic=True, color="AAAAAA")
    c.alignment = aln("center","center")

    # Header
    ws.row_dimensions[3].height = 42
    hdrs = [("Name",C_HEADER_TAGE),(TAGE_HEADER[0],C_HEADER_TAGE),(TAGE_HEADER[1],C_HEADER_TAGE),
            (TAGE_HEADER[2],C_HEADER_TAGE),(TAGE_HEADER[3],C_HEADER_TAGE),
            ("Ergebnis",C_HEADER_ERG),("Insgesamt\nPunkte",C_HEADER_INSG)]
    for i,(h,bg) in enumerate(hdrs):
        c = ws.cell(row=3,column=i+1)
        c.value=h; c.fill=fill(bg); c.font=fnt(11,bold=True,color=C_WHITE)
        c.alignment=aln("center","center",wrap=True); c.border=thin()

    ws.freeze_panes = "B4"

    cur = 4; alt = True; last_data = 4
    for gk, mitglieder in GRUPPEN.items():
        if cur > 4:
            ws.row_dimensions[cur].height = 5
            ws.merge_cells(f"A{cur}:G{cur}")
            ws[f"A{cur}"].fill = fill(C_SEPARATOR)
            cur += 1
        ws.row_dimensions[cur].height = 22
        ws.merge_cells(f"A{cur}:G{cur}")
        c = ws[f"A{cur}"]
        c.value=f"  ⚡  Gruppe {gk[-1]}"; c.fill=fill(C_SUBTITLE_BG)
        c.font=fnt(11,bold=True,color=C_WHITE); c.alignment=aln("left","center"); c.border=medium()
        cur += 1; alt = True

        for name, tage in mitglieder.items():
            rbg = C_ROW_LIGHT if alt else C_ROW_GREY; alt = not alt
            ws.row_dimensions[cur].height = 22
            rf = fill(rbg)
            # A: Name
            c = ws.cell(row=cur,column=1)
            c.value=f"  {name}"; c.fill=rf
            c.font=fnt(11,color=("FF0000" if name.split()[0] in ABWESEND else "1A1A2E"))
            c.alignment=aln("left","center"); c.border=thin()
            # B-E: Tage
            for t in range(4):
                c=ws.cell(row=cur,column=t+2)
                if t in tage:
                    c.value="✓"; c.fill=fill(C_CHECK_BG); c.font=fnt(14,bold=True,color=C_CHECK_FG)
                else: c.fill=rf
                c.alignment=aln("center","center"); c.border=thin()
            # F: Ergebnis
            c=ws.cell(row=cur,column=6)
            c.value=f"=COUNTA(B{cur}:E{cur})"
            c.font=fnt(13,bold=True,color=C_WHITE); c.alignment=aln("center","center"); c.border=thin()
            # G: Insgesamt
            base=get_vor(gk,name)-get_reset(gk,name)
            c=ws.cell(row=cur,column=7)
            c.value=f"=MAX(0,{base}+COUNTA(B{cur}:E{cur}))"
            c.fill=fill(C_INSG_BG); c.font=fnt(13,bold=True,color="1A1A2E")
            c.alignment=aln("center","center"); c.border=thin()
            last_data=cur; cur+=1

    # Legende
    ws.row_dimensions[cur].height=5; ws.merge_cells(f"A{cur}:G{cur}")
    ws[f"A{cur}"].fill=fill(C_SEPARATOR); cur+=1
    ws.row_dimensions[cur].height=18; ws.merge_cells(f"A{cur}:G{cur}")
    c=ws[f"A{cur}"]
    c.value="  Legende: ✓ = trainiert | 0–1 \U0001F534 | 2 \U0001F7E0 | 3–4 \U0001F7E2 | Insgesamt = aktuell verdient"
    c.fill=fill(C_LEGEND_BG); c.font=fnt(9,italic=True,color="555555"); c.alignment=aln("left","center")

    # Conditional Formatting F-Spalte
    dr=f"F4:F{last_data}"
    ws.conditional_formatting.add(dr,CellIsRule("between",["0","1"],
        fill=PatternFill("solid",fgColor=C_RED),font=Font(name="Arial",color=C_WHITE,bold=True)))
    ws.conditional_formatting.add(dr,CellIsRule("equal",["2"],
        fill=PatternFill("solid",fgColor=C_ORANGE),font=Font(name="Arial",color=C_WHITE,bold=True)))
    ws.conditional_formatting.add(dr,CellIsRule("between",["3","4"],
        fill=PatternFill("solid",fgColor=C_GREEN),font=Font(name="Arial",color=C_WHITE,bold=True)))

    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_setup.fitToHeight=1; ws.page_setup.fitToWidth=1
    wb.save(XLSX_PATH)
    print(f"OK XLSX: {XLSX_PATH}")

# ===================================================================
#  PDF ERSTELLEN (reportlab)
# ===================================================================

def create_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib.units import mm

    def hx(s): return rl.Color(int(s[0:2],16)/255,int(s[2:4],16)/255,int(s[4:6],16)/255)

    CW=[58*mm,20*mm,20*mm,20*mm,20*mm,20*mm,25*mm]
    data=[]; st=[]; r=0

    # Titel
    sd=f"{START_DATUM[0:2]}.{START_DATUM[3:5]}."; ed=f"{END_DATUM[0:2]}.{END_DATUM[3:5]}.{END_DATUM[6:10]}"
    data.append([f"WOCHENCHALLENGE - {sd} - {ed}","","","","","",""])
    st+=[ ("SPAN",(0,r),(6,r)),("BACKGROUND",(0,r),(6,r),hx(C_TITLE_BG)),
          ("TEXTCOLOR",(0,r),(6,r),rl.white),("FONTNAME",(0,r),(6,r),"Helvetica-Bold"),
          ("FONTSIZE",(0,r),(6,r),13),("ALIGN",(0,r),(6,r),"CENTER"),("VALIGN",(0,r),(6,r),"MIDDLE")]
    r+=1

    data.append(["1 Trainingstag = 1 Punkt | Max. 4 Punkte | 0-1 ROT  2 ORANGE  3-4 GRUEN","","","","","",""])
    st+=[ ("SPAN",(0,r),(6,r)),("BACKGROUND",(0,r),(6,r),hx(C_SUBTITLE_BG)),
          ("TEXTCOLOR",(0,r),(6,r),rl.HexColor("#AAAAAA")),("FONTNAME",(0,r),(6,r),"Helvetica-Oblique"),
          ("FONTSIZE",(0,r),(6,r),8),("ALIGN",(0,r),(6,r),"CENTER"),("VALIGN",(0,r),(6,r),"MIDDLE")]
    r+=1

    data.append(["Name"] + TAGE_HEADER + ["Ergebnis", "Insgesamt"])
    st+=[ ("BACKGROUND",(0,r),(4,r),hx(C_HEADER_TAGE)),
          ("BACKGROUND",(5,r),(5,r),hx(C_HEADER_ERG)),
          ("BACKGROUND",(6,r),(6,r),hx(C_HEADER_INSG)),
          ("TEXTCOLOR",(0,r),(6,r),rl.white),("FONTNAME",(0,r),(6,r),"Helvetica-Bold"),
          ("FONTSIZE",(0,r),(6,r),9),("ALIGN",(0,r),(6,r),"CENTER"),("VALIGN",(0,r),(6,r),"MIDDLE")]
    r+=1

    alt=True
    for gk,mitglieder in GRUPPEN.items():
        data.append(["","","","","","",""])
        st+=[ ("BACKGROUND",(0,r),(6,r),hx(C_SEPARATOR)),
              ("TOPPADDING",(0,r),(6,r),0),("BOTTOMPADDING",(0,r),(6,r),0)]
        r+=1
        data.append([f"  Gruppe {gk[-1]}","","","","","",""])
        st+=[ ("SPAN",(0,r),(6,r)),("BACKGROUND",(0,r),(6,r),hx(C_SUBTITLE_BG)),
              ("TEXTCOLOR",(0,r),(6,r),rl.white),("FONTNAME",(0,r),(6,r),"Helvetica-Bold"),
              ("FONTSIZE",(0,r),(6,r),10),("ALIGN",(0,r),(6,r),"LEFT"),
              ("VALIGN",(0,r),(6,r),"MIDDLE"),("LEFTPADDING",(0,r),(6,r),8)]
        r+=1; alt=True

        for name,tage in mitglieder.items():
            punkte=len(tage); ins=max(0, get_vor(gk,name)-get_reset(gk,name)+punkte)
            bg=hx(C_ROW_LIGHT) if alt else hx(C_ROW_GREY); alt=not alt
            row=[f"  {name}"]+["✓" if t in tage else "" for t in range(4)]+[str(punkte),str(ins)]
            data.append(row)
            st+=[ ("BACKGROUND",(0,r),(6,r),bg),("FONTNAME",(0,r),(0,r),"Helvetica"),
                  ("FONTSIZE",(0,r),(6,r),9),("ALIGN",(0,r),(0,r),"LEFT"),
                  ("ALIGN",(1,r),(6,r),"CENTER"),("VALIGN",(0,r),(6,r),"MIDDLE"),
                  ("BACKGROUND",(6,r),(6,r),hx(C_INSG_BG)),
                  ("FONTNAME",(6,r),(6,r),"Helvetica-Bold")]
            if name.split()[0] in ABWESEND:
                st.append(("TEXTCOLOR",(0,r),(0,r),rl.HexColor("#FF0000")))
            for t in range(4):
                if t in tage:
                    st+=[ ("BACKGROUND",(t+1,r),(t+1,r),hx(C_CHECK_BG)),
                          ("TEXTCOLOR",(t+1,r),(t+1,r),hx(C_CHECK_FG)),
                          ("FONTNAME",(t+1,r),(t+1,r),"Helvetica-Bold"),
                          ("FONTSIZE",(t+1,r),(t+1,r),11)]
            ebg=hx(C_RED) if punkte<=1 else (hx(C_ORANGE) if punkte==2 else hx(C_GREEN))
            st+=[ ("BACKGROUND",(5,r),(5,r),ebg),("TEXTCOLOR",(5,r),(5,r),rl.white),
                  ("FONTNAME",(5,r),(5,r),"Helvetica-Bold"),("FONTSIZE",(5,r),(5,r),10)]
            r+=1

    st+=[ ("GRID",(0,0),(-1,-1),0.4,rl.HexColor("#CCCCCC")),
          ("BOX",(0,0),(-1,-1),1.2,rl.HexColor("#1A1A2E"))]

    rh=[26 if i==0 else 14 if i==1 else 28 if i==2 else 4 if all(d=="" for d in row) else 18
        for i,row in enumerate(data)]

    t=Table(data,colWidths=CW,rowHeights=rh)
    t.setStyle(TableStyle(st))
    doc=SimpleDocTemplate(PDF_PATH,pagesize=A4,
        leftMargin=12*mm,rightMargin=12*mm,topMargin=12*mm,bottomMargin=12*mm)
    doc.build([t])
    print(f"OK PDF:  {PDF_PATH}")

# ===================================================================
#  MAIN
# ===================================================================

if __name__ == "__main__":
    print(f"=== Wochenchallenge {START_DATUM} - {END_DATUM} ===")
    create_xlsx()
    create_pdf()
    print("=== Fertig ===")
